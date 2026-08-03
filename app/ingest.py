"""Parse a Slate .xlsx export and merge it into the database.

Stage flags come from analysis_engine.ft_stages / summer_stages, evaluated
against the RAW openpyxl row (dates as datetimes, Y/N as-is) so the canonical
logic sees exactly what it expects. Normalisation for storage happens after.
"""
import datetime as _dt
import hashlib
import io
import zipfile
from collections import defaultdict

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from . import programs
from .analysis_engine import build_path, classify

SHEET = "Export"


class IngestError(Exception):
    pass


def _s(v):
    """Any cell -> trimmed string ('' for None)."""
    if v is None:
        return ""
    if isinstance(v, (_dt.datetime, _dt.date)):
        return _iso(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _iso(v):
    """Date-ish cell -> 'YYYY-MM-DD' ('' when absent/unparseable)."""
    if v is None:
        return ""
    if isinstance(v, _dt.datetime):
        return v.date().isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    text = str(v).strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return _dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return text[:10]


def _age(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def read_rows(path_or_stream):
    """-> (headers, data_rows). Raises IngestError on a shape we can't read."""
    try:
        wb = openpyxl.load_workbook(path_or_stream, read_only=True, data_only=True)
    except Exception as exc:
        raise IngestError("Could not open as .xlsx: %s" % exc)
    if SHEET in wb.sheetnames:
        ws = wb[SHEET]
    elif len(wb.sheetnames) == 1:
        ws = wb[wb.sheetnames[0]]
    else:
        raise IngestError(
            "Expected a sheet named %r; found %s" % (SHEET, wb.sheetnames)
        )
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    if not rows:
        raise IngestError("File is empty.")
    headers = rows[0]
    data = [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]
    if not data:
        raise IngestError("File has a header row but no data rows.")
    return headers, data


def validate(program, headers, data):
    """Loose structural check. Returns (layout, warnings).

    Slate's arrangement changes over time, so the file is matched against every
    known Layout rather than one fixed map -- see programs.Layout. A file that
    matches none of them is refused outright: importing anyway would silently
    mis-read every field after the shift.
    """
    warnings = []
    if len(headers) < program.min_cols:
        raise IngestError(
            "%s exports need at least %d columns; this file has %d. "
            "Wrong program selected?" % (program.label, program.min_cols, len(headers))
        )

    layout, shifted = program.layout_for(headers)
    if layout is None:
        detail = "; ".join(
            "column %d should contain %r but is %r" % (i, e, f) for i, e, f in shifted)
        raise IngestError(
            "This file's columns match none of the %d known %s layouts. Closest is "
            "%r: %s. Nothing was imported. If Slate changed the export again, add "
            "the new arrangement as a Layout in app/programs.py — importing anyway "
            "would silently mis-read every field after the shift."
            % (len(program.layouts), program.label, program.layouts[0].name, detail)
        )
    if layout is not program.layouts[0]:
        warnings.append(
            "Read using the older %r column layout. That is expected for an export "
            "pulled before Slate's latest change." % layout.name
        )

    norm = " | ".join(_s(h).lower() for h in headers)
    missing = [h for h in programs.EXPECTED_HEADER_HINTS[program.key] if h not in norm]
    if missing:
        warnings.append(
            "Header check: expected to see %s in the header row. Columns are read "
            "by position, so this is a heads-up, not a failure." % ", ".join(repr(m) for m in missing)
        )
    extra = len(headers) - layout.min_cols
    if extra > 0:
        warnings.append(
            "%d column(s) beyond the %d this layout maps were ignored. Appended "
            "columns are safe; if one of them is meaningful (a new program type, "
            "say) it needs adding to app/programs.py to be used."
            % (extra, layout.min_cols)
        )
    short = sum(1 for r in data if len(r) < layout.min_cols)
    if short:
        warnings.append("%d row(s) are shorter than expected and were padded." % short)
    return layout, warnings


def _pad(row, n):
    return row if len(row) >= n else tuple(row) + (None,) * (n - len(row))


def extract_pings(row, utm_idx):
    """Raw row -> list of ping dicts with ts/seq/source/medium/campaign/content
    plus the canonical channel + sub_source. Order is chronological."""
    si, mi, ci, ti = utm_idx
    out = []
    untimed = 0
    for p in build_path(row[si], row[mi], row[ci], row[ti]):
        ts = p.get("ts", "") or ""
        if ts:
            seq = 0
        else:
            untimed += 1
            seq = untimed
        source, medium = p.get("s"), p.get("m")
        campaign, content = p.get("c"), p.get("t")
        channel, sub = classify(source, medium, campaign)
        if channel is None:
            continue  # no source AND no medium -> not a ping we can attribute
        out.append({
            "ts": ts, "seq": seq,
            "source": _s(source), "medium": _s(medium),
            "campaign": _s(campaign), "content": _s(content),
            "channel": channel, "sub_source": sub or "",
        })
    return out


def _check_term_shrinkage(conn, program_key, layout, data, pad_to):
    """Flag any term where this file has FEWER rows than the database already
    holds for it.

    Every export is expected to be a superset of the last one -- cumulative,
    never narrower. A term coming back smaller almost always means the export
    itself was scoped differently (a term or date filter on Slate's side), not
    that people vanished; the July -> August pull dropped 1,326 of 1,334 Winter
    2026 rows exactly this way. It is a warning, not a hard stop: uploads merge
    and never delete, so accepting the file is always safe, but silently eating
    a shrunk term is how the wrong Slate report goes unnoticed for months.
    """
    term_idx = layout.cols.get("term")
    if term_idx is None:
        return []

    have = {r["term"]: r["n"] for r in conn.execute(
        "SELECT term, COUNT(*) n FROM applicants WHERE program=? GROUP BY term",
        (program_key,))}
    if not have:
        return []          # first upload ever for this program -- nothing to compare

    incoming = defaultdict(int)
    for raw in data:
        row = _pad(raw, pad_to)
        incoming[_s(row[term_idx])] += 1

    shrunk = sorted(
        ((term, before, incoming.get(term, 0)) for term, before in have.items()
         if incoming.get(term, 0) < before),
        key=lambda t: t[1] - t[2], reverse=True)
    if not shrunk:
        return []

    lines = ["%r: had %d, this file has %d" % (t or "(blank)", b, n)
             for t, b, n in shrunk[:6]]
    if len(shrunk) > 6:
        lines.append("and %d more term(s)" % (len(shrunk) - 6))
    return ["This file has FEWER rows than already stored for %d term(s) — "
            "%s. Nobody was deleted (uploads only ever add or update), but this "
            "usually means the export itself was scoped narrower than before "
            "(a term or date filter on Slate's side) rather than that people "
            "left. Worth checking before trusting counts for those terms."
            % (len(shrunk), "; ".join(lines))]


def ingest(conn, path_or_stream, program_key, filename, sha256=None):
    """Merge one export into the DB. Returns a summary dict."""
    program = programs.get(program_key)
    headers, data = read_rows(path_or_stream)
    layout, warnings = validate(program, headers, data)
    # A dedicated enrolled column is expected in a future export; until then the
    # stage is read off Most Recent Decision. Detected per file, so one export
    # with the column and one without both ingest correctly.
    enrolled_idx = programs.enrolled_index(headers)
    if enrolled_idx is not None and "enrolled" in program.stage_keys:
        warnings.append(
            "Reading Enrolled from column %d (%r) instead of Most Recent Decision."
            % (enrolled_idx + 1, programs._s(headers[enrolled_idx])))
    pad_to = max(layout.min_cols, (enrolled_idx or -1) + 1)
    warnings.extend(_check_term_shrinkage(conn, program_key, layout, data, pad_to))

    if sha256 is None:
        sha256 = ""
    dup_file = None
    if sha256:
        prev = conn.execute(
            "SELECT id, filename, uploaded_at FROM uploads WHERE sha256=? AND program=?",
            (sha256, program_key),
        ).fetchone()
        if prev:
            dup_file = dict(prev)
            warnings.append(
                "This exact file was already uploaded on %s as %r. Re-ingesting is "
                "harmless (everything de-dupes), but it will add no new data."
                % (dup_file["uploaded_at"], dup_file["filename"])
            )

    cur = conn.cursor()
    cur.execute(
        "INSERT INTO uploads (program, filename, sha256, uploaded_at, row_count) "
        "VALUES (?,?,?,?,?)",
        (program_key, filename, sha256,
         _dt.datetime.now().replace(microsecond=0).isoformat(" "), len(data)),
    )
    upload_id = cur.lastrowid

    cols = layout.cols
    stage_map = {
        "started": "st_started", "submitted": "st_submitted",
        "aud_req": "st_aud_req", "aud_comp": "st_aud_comp",
        "admitted": "st_admitted", "accepted": "st_accepted",
        "enrolled": "st_enrolled",
    }

    n_new = n_upd = pings_new = pings_dup = 0
    unknown = defaultdict(lambda: {"pings": 0, "applicants": set()})
    skipped_no_id = 0
    # Disambiguates rows that collide on (global_id, term, started_date) within
    # a single file. Zero for both sample files; exists so a future collision
    # cannot silently drop a row.
    seen_keys = defaultdict(int)
    repeated_ids = defaultdict(int)

    for raw in data:
        row = _pad(raw, pad_to)
        global_id = _s(row[cols["global_id"]])
        if not global_id:
            skipped_no_id += 1
            continue
        repeated_ids[global_id] += 1

        # Canonical engine for the stages it owns, plus any derived here.
        flags = program.stages(row, enrolled_idx)
        stage_vals = {v: 0 for v in stage_map.values()}
        for k, v in flags.items():
            stage_vals[stage_map[k]] = 1 if v else 0

        fields = {
            "term": _s(row[cols["term"]]),
            "country": _s(row[cols.get("country", 0)]) if "country" in cols else "",
            "region": _s(row[cols["region"]]) if "region" in cols else "",
            "city": _s(row[cols["city"]]) if "city" in cols else "",
            "postal": _s(row[cols["postal"]]) if "postal" in cols else "",
            "age": _age(row[cols["age"]]) if "age" in cols else None,
            "emphasis": _s(row[cols["emphasis"]]) if "emphasis" in cols else "",
            "decision": _s(row[cols["decision"]]) if "decision" in cols else "",
            "app_status": _s(row[cols["app_status"]]) if "app_status" in cols else "",
            "started_date": _iso(row[cols["started_date"]]) if "started_date" in cols else "",
            "submitted_date": _iso(row[cols["submitted_date"]]) if "submitted_date" in cols else "",
            "completed_date": _iso(row[cols["completed_date"]]) if "completed_date" in cols else "",
        }

        # Dedup identity: (global_id, term, started_date) + an occurrence
        # ordinal. Term and start date never change once an application exists,
        # so this key is stable from export to export.
        ident = (global_id, fields["term"], fields["started_date"])
        dedup_seq = seen_keys[ident]
        seen_keys[ident] += 1

        existing = cur.execute(
            "SELECT id FROM applicants WHERE program=? AND global_id=? AND term=?"
            " AND started_date=? AND dedup_seq=?",
            (program_key, global_id, fields["term"], fields["started_date"], dedup_seq),
        ).fetchone()

        if existing is None:
            cur.execute(
                "INSERT INTO applicants (program, global_id, term, country, region, city,"
                " postal, age, emphasis, decision, app_status, started_date,"
                " submitted_date, completed_date, st_started, st_submitted, st_aud_req,"
                " st_aud_comp, st_admitted, st_accepted, st_enrolled, dedup_seq,"
                " first_upload_id,"
                " last_upload_id)"
                " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (program_key, global_id, fields["term"], fields["country"],
                 fields["region"], fields["city"], fields["postal"], fields["age"],
                 fields["emphasis"], fields["decision"], fields["app_status"],
                 fields["started_date"], fields["submitted_date"],
                 fields["completed_date"],
                 stage_vals["st_started"], stage_vals["st_submitted"],
                 stage_vals["st_aud_req"], stage_vals["st_aud_comp"],
                 stage_vals["st_admitted"], stage_vals["st_accepted"],
                 stage_vals["st_enrolled"],
                 dedup_seq, upload_id, upload_id),
            )
            applicant_id = cur.lastrowid
            n_new += 1
        else:
            applicant_id = existing["id"]
            # Descriptive fields: newer non-empty value wins.
            # Stage flags: monotonic MAX so ingest order cannot regress a stage.
            cur.execute(
                "UPDATE applicants SET"
                "  term=CASE WHEN ?<>'' THEN ? ELSE term END,"
                "  country=CASE WHEN ?<>'' THEN ? ELSE country END,"
                "  region=CASE WHEN ?<>'' THEN ? ELSE region END,"
                "  city=CASE WHEN ?<>'' THEN ? ELSE city END,"
                "  postal=CASE WHEN ?<>'' THEN ? ELSE postal END,"
                "  age=COALESCE(?, age),"
                "  emphasis=CASE WHEN ?<>'' THEN ? ELSE emphasis END,"
                "  decision=CASE WHEN ?<>'' THEN ? ELSE decision END,"
                "  app_status=CASE WHEN ?<>'' THEN ? ELSE app_status END,"
                "  started_date=CASE WHEN ?<>'' THEN ? ELSE started_date END,"
                "  submitted_date=CASE WHEN ?<>'' THEN ? ELSE submitted_date END,"
                "  completed_date=CASE WHEN ?<>'' THEN ? ELSE completed_date END,"
                "  st_started=MAX(st_started,?), st_submitted=MAX(st_submitted,?),"
                "  st_aud_req=MAX(st_aud_req,?), st_aud_comp=MAX(st_aud_comp,?),"
                "  st_admitted=MAX(st_admitted,?), st_accepted=MAX(st_accepted,?),"
                "  st_enrolled=MAX(st_enrolled,?),"
                "  last_upload_id=?"
                " WHERE id=?",
                (fields["term"], fields["term"], fields["country"], fields["country"],
                 fields["region"], fields["region"], fields["city"], fields["city"],
                 fields["postal"], fields["postal"], fields["age"],
                 fields["emphasis"], fields["emphasis"],
                 fields["decision"], fields["decision"],
                 fields["app_status"], fields["app_status"],
                 fields["started_date"], fields["started_date"],
                 fields["submitted_date"], fields["submitted_date"],
                 fields["completed_date"], fields["completed_date"],
                 stage_vals["st_started"], stage_vals["st_submitted"],
                 stage_vals["st_aud_req"], stage_vals["st_aud_comp"],
                 stage_vals["st_admitted"], stage_vals["st_accepted"],
                 stage_vals["st_enrolled"],
                 upload_id, applicant_id),
            )
            n_upd += 1

        for p in extract_pings(row, layout.utm_idx):
            cur.execute(
                "INSERT OR IGNORE INTO pings (applicant_id, ts, seq, source, medium,"
                " campaign, content, channel, sub_source) VALUES (?,?,?,?,?,?,?,?,?)",
                (applicant_id, p["ts"], p["seq"], p["source"], p["medium"],
                 p["campaign"], p["content"], p["channel"], p["sub_source"]),
            )
            if cur.rowcount:
                pings_new += 1
            else:
                pings_dup += 1
            if p["channel"] == "Unresolved/Other":
                k = (p["source"], p["medium"], p["campaign"])
                unknown[k]["pings"] += 1
                unknown[k]["applicants"].add(applicant_id)

    now = _dt.datetime.now().replace(microsecond=0).isoformat(" ")
    # Recompute the review queue from the deduplicated `pings` table rather than
    # incrementing counters. Incrementing double-counts on re-upload: the pings
    # themselves de-duplicate, so a second copy of the same export must not make
    # an unclassified combo look more common than it is.
    # `acknowledged` is only cleared when the count actually grew, so re-ingesting
    # an old export doesn't un-review something already dealt with.
    cur.execute(
        "INSERT INTO unknown_utms (source, medium, campaign, ping_count,"
        " applicant_count, first_upload_id, last_seen_at)"
        " SELECT source, medium, campaign, COUNT(*), COUNT(DISTINCT applicant_id),"
        "        ?, ?"
        "   FROM pings WHERE channel = 'Unresolved/Other'"
        "  GROUP BY source, medium, campaign"
        " ON CONFLICT(source, medium, campaign) DO UPDATE SET"
        "   acknowledged=CASE WHEN excluded.ping_count > unknown_utms.ping_count"
        "                     THEN 0 ELSE unknown_utms.acknowledged END,"
        "   ping_count=excluded.ping_count,"
        "   applicant_count=excluded.applicant_count,"
        "   last_seen_at=excluded.last_seen_at",
        (upload_id, now),
    )

    if skipped_no_id:
        warnings.append("%d row(s) had no Global ID and were skipped." % skipped_no_id)

    repeats = {g: c for g, c in repeated_ids.items() if c > 1}
    if repeats:
        warnings.append(
            "%d Global ID(s) appear on more than one row (%d extra rows). Each row is "
            "kept as its own record, keyed on Global ID + term + start date, which is "
            "what the reference analysis counts. Usually one row carries a term and a "
            "duplicate carries a blank term."
            % (len(repeats), sum(repeats.values()) - len(repeats))
        )
    collisions = sum(v - 1 for v in seen_keys.values() if v > 1)
    if collisions:
        warnings.append(
            "%d row(s) matched an existing row on Global ID + term + start date and "
            "were stored with an occurrence number to avoid dropping them. Worth a "
            "look -- this is unusual." % collisions
        )

    cur.execute(
        "UPDATE uploads SET applicants_new=?, applicants_updated=?, pings_new=?,"
        " pings_duplicate=?, notes=? WHERE id=?",
        (n_new, n_upd, pings_new, pings_dup, "\n".join(warnings), upload_id),
    )
    conn.commit()

    return {
        "upload_id": upload_id, "program": program_key, "filename": filename,
        "row_count": len(data), "applicants_new": n_new,
        "applicants_updated": n_upd, "pings_new": pings_new,
        "pings_duplicate": pings_dup, "warnings": warnings,
        "duplicate_file": dup_file,
        "new_unknown_utms": [
            {"source": s, "medium": m, "campaign": c, "pings": a["pings"],
             "applicants": len(a["applicants"])}
            for (s, m, c), a in sorted(unknown.items(), key=lambda kv: -kv[1]["pings"])
        ],
    }


def sha256_of(data_bytes):
    return hashlib.sha256(data_bytes).hexdigest()


# ---------------------------------------------------------------------------
# Slate native "referrer" ping log — DATA_SCHEMA.md's optional third export.
# One row per PAGE VIEW. Its value is catching untagged traffic the UTM feed
# cannot see (email opens via the Gmail app, Zoom sessions, AI referrals,
# partner sites), so the referrer domain is the thing worth reporting.
# ---------------------------------------------------------------------------
OWN_DOMAINS = ("aada.edu",)   # anything under these is our own site, not a source

REF_ANCHORS = {
    0: "reference id", 1: "referrer", 2: "duration", 3: "timestamp", 4: "url",
    5: "campaign", 6: "medium", 7: "content", 8: "source", 9: "term",
}


def ref_domain(url):
    """Referrer URL -> bare host ('' when absent/unparseable)."""
    from urllib.parse import urlsplit
    if not url:
        return ""
    try:
        host = (urlsplit(str(url)).netloc or "").lower()
    except ValueError:
        return ""
    if not host:
        # Android app referrers arrive bare, e.g. "com.google.android.gm".
        bare = str(url).strip().lower()
        return bare if bare and "/" not in bare and " " not in bare else ""
    if host.startswith("www."):
        host = host[4:]
    return host.split(":")[0]


def is_internal(domain):
    return bool(domain) and any(
        domain == d or domain.endswith("." + d) for d in OWN_DOMAINS)


def ingest_referrals(conn, path_or_stream, filename, sha256=None):
    """Merge one native referrer ping log. Returns a summary dict."""
    headers, data = read_rows(path_or_stream)

    shifted = []
    for idx, expect in sorted(REF_ANCHORS.items()):
        found = _s(headers[idx]).lower() if idx < len(headers) else ""
        if expect not in found:
            shifted.append((idx, expect, _s(headers[idx]) if idx < len(headers) else "(missing)"))
    if shifted:
        raise IngestError(
            "This does not look like a Slate referrer ping log: %s. Expected the "
            "10 columns in DATA_SCHEMA.md (Person Reference ID, Ping Referrer, "
            "Ping Duration, Ping Timestamp, Ping URL, then Ping UTM "
            "Campaign/Medium/Content/Source/Term)."
            % "; ".join("column %d should contain %r but is %r" % s for s in shifted)
        )

    warnings = []
    dup_file = None
    if sha256:
        prev = conn.execute(
            "SELECT id, filename, uploaded_at FROM ref_uploads WHERE sha256=?",
            (sha256,)).fetchone()
        if prev:
            dup_file = dict(prev)
            warnings.append(
                "This exact file was already uploaded on %s as %r. Re-ingesting is "
                "harmless (page views de-duplicate) but adds nothing new."
                % (dup_file["uploaded_at"], dup_file["filename"]))

    cur = conn.cursor()
    cur.execute(
        "INSERT INTO ref_uploads (filename, sha256, uploaded_at, row_count)"
        " VALUES (?,?,?,?)",
        (filename, sha256 or "",
         _dt.datetime.now().replace(microsecond=0).isoformat(" "), len(data)))
    upload_id = cur.lastrowid

    people = set()
    skipped = 0
    batch = []
    new = dup = 0

    def flush():
        nonlocal new, dup
        if not batch:
            return
        before = conn.total_changes
        cur.executemany(
            "INSERT OR IGNORE INTO ref_pings (global_id, ts, duration, referrer,"
            " domain, internal, url, source, medium, campaign, content, term,"
            " has_utm) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", batch)
        inserted = conn.total_changes - before
        new += inserted
        dup += len(batch) - inserted
        del batch[:]

    for raw in data:
        row = _pad(raw, 10)
        gid = _s(row[0])
        if not gid:
            skipped += 1
            continue
        people.add(gid)
        referrer = _s(row[1])
        domain = ref_domain(referrer)
        ts = row[3]
        if isinstance(ts, (_dt.datetime, _dt.date)):
            ts = ts.isoformat(" ") if isinstance(ts, _dt.datetime) else ts.isoformat()
        else:
            ts = _s(ts)
        try:
            duration = int(float(row[2])) if row[2] is not None else None
        except (TypeError, ValueError):
            duration = None
        source, medium = _s(row[8]), _s(row[6])
        batch.append((
            gid, ts, duration, referrer, domain, 1 if is_internal(domain) else 0,
            _s(row[4]), source, medium, _s(row[5]), _s(row[7]), _s(row[9]),
            1 if (source or medium) else 0,
        ))
        if len(batch) >= 4000:
            flush()
    flush()

    if skipped:
        warnings.append("%d row(s) had no Person Reference ID and were skipped." % skipped)

    cur.execute(
        "UPDATE ref_uploads SET pings_new=?, pings_duplicate=?, people=?, notes=?"
        " WHERE id=?", (new, dup, len(people), "\n".join(warnings), upload_id))
    conn.commit()

    return {"upload_id": upload_id, "filename": filename, "row_count": len(data),
            "pings_new": new, "pings_duplicate": dup, "people": len(people),
            "warnings": warnings, "duplicate_file": dup_file}

# --------------------------------------------------------------------------
# what kind of file is this?
# --------------------------------------------------------------------------
# Upload kinds, matching the values the /uploads form posts.
KIND_FT = "ft"
KIND_SUMMER = "summer"
KIND_REFERRALS = "referrals"
KIND_GOOGLE = "spend_google"
KIND_META = "spend_meta"


def sniff_kind(filename, payload):
    """Guess what an uploaded file is -> (kind, confidence, why).

    Headers first, filename only as a tie-breaker: a file's contents are what it
    actually IS, whereas its name is whatever someone typed. Every guess is
    shown to the user before anything is imported, and is overridable -- this
    saves picking from a dropdown five times, it does not get to be wrong
    silently.

    confidence: "sure" (header match) or "guess" (filename only).
    """
    name = (filename or "").lower()

    if name.endswith(".csv"):
        head = ""
        try:
            head = payload[:4000].decode("utf-8-sig", "replace").lower()
        except Exception:
            pass
        if "campaign" in head and "cost" in head:
            return KIND_GOOGLE, "sure", "CSV with Campaign and Cost columns"
        return KIND_GOOGLE, "guess", "CSV -- Google Ads is the only CSV type"

    headers = []
    try:
        import warnings as _w
        from openpyxl import load_workbook
        with _w.catch_warnings():
            _w.simplefilter("ignore")
            wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [_s(h).lower() for h in row]
                break
            sheets = [n.lower() for n in wb.sheetnames]
            wb.close()
    except (OSError, ValueError, KeyError, TypeError,
            zipfile.BadZipFile, InvalidFileException):
        # Genuinely unreadable as a workbook -- fall back to the filename.
        # Deliberately NOT a bare `except Exception`: a missing import once hid
        # here and silently downgraded every xlsx to a filename guess.
        return _sniff_by_name(name)

    joined = " | ".join(headers)

    if "amount spent" in joined or "ad set name" in joined:
        return KIND_META, "sure", "Meta export (Amount spent column)"
    if "campaign" in joined and "cost" in joined:
        return KIND_GOOGLE, "sure", "Google Ads export (Campaign + Cost)"

    # Slate exports all carry Global ID; the program is the question.
    if "global id" in joined:
        if "full-time app term" in joined or "started ft app date" in joined:
            return KIND_FT, "sure", "Slate export with Full-Time columns"
        if "app date" in joined and "application status" in joined:
            return KIND_SUMMER, "sure", "Slate export with Summer columns"
        if "referrer" in joined or "url" in joined:
            return KIND_REFERRALS, "sure", "Slate referrer ping log"
        guess, _c, _w2 = _sniff_by_name(name)
        return guess, "guess", "Slate export; program taken from the filename"

    if any("referral" in x or "referrer" in x for x in sheets + headers):
        return KIND_REFERRALS, "sure", "referrer ping log"

    return _sniff_by_name(name)


def _sniff_by_name(name):
    if "summer" in name:
        return KIND_SUMMER, "guess", "filename says Summer"
    if "referral" in name or "referrer" in name:
        return KIND_REFERRALS, "guess", "filename says referral"
    if "meta" in name:
        return KIND_META, "guess", "filename says Meta"
    if "google" in name:
        return KIND_GOOGLE, "guess", "filename says Google"
    if "2 year" in name or "full" in name or "ping data" in name:
        return KIND_FT, "guess", "filename looks like a Full-Time ping export"
    return KIND_FT, "guess", "defaulted to Full-Time"
