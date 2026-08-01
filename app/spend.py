"""Media spend: parse Google Ads and Meta exports onto the app's own taxonomy.

Why there is no campaign mapping table
--------------------------------------
The obvious design is a lookup from platform campaign name to channel. It isn't
needed. Google's export carries `Campaign type` (Performance Max / Search /
Display) and Meta's carries `Platform` (instagram / facebook), and those land
exactly on the sub-sources the app already charts. Google's *utm_campaign*
values are hand-made codes ("PMaxInterest", "SearchUS") that do not resemble its
campaign names, so a campaign-level join WOULD need a hand-maintained mapping —
rolling up by campaign type avoids inventing one while still resolving to the
granularity every chart here draws.

Columns are matched BY HEADER NAME, not by position. Slate exports get the
opposite treatment (see programs.HEADER_ANCHORS) because they are a fixed
machine-generated layout where a shifted column is silent corruption. These are
user-configured platform reports: the column order changes with report settings,
and a missing column should say so rather than read the one next to it.
"""
import csv
import datetime as dt
import io
import re

from . import taxonomy

GOOGLE = "google"
META = "meta"
PLATFORM_LABELS = {GOOGLE: "Google Ads", META: "Meta Ads"}

GOOGLE_CHANNEL = "Google (Paid)"
META_CHANNEL = "Meta Paid Social (IG/FB)"

# Campaign type -> the sub-source the app already uses for it.
GOOGLE_TYPE_SUB = {
    "performance max": "PMax",
    "search": "Search",
    "display": "Other paid",
    "video": "Other paid",
    "demand gen": "Other paid",
    "discovery": "Other paid",
    "shopping": "Other paid",
    "app": "Other paid",
}
GOOGLE_FALLBACK_SUB = "Other paid"

META_PLATFORM_SUB = {
    "instagram": "Instagram",
    "facebook": "Facebook",
    "audience_network": "Other Meta",
    "messenger": "Other Meta",
    "threads": "Other Meta",
    "unknown": "Other Meta",
}
META_FALLBACK_SUB = "Other Meta"

MONTH_NAMES = {m.lower(): i for i, m in enumerate(
    ["", "January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"])}


class SpendError(Exception):
    """A file we cannot read. Surfaced to the user, never swallowed."""


def _s(v):
    return "" if v is None else str(v).strip()


def _num(v):
    """Money/counts out of a platform export: '1,234.56' -> 1234.56."""
    t = _s(v).replace(",", "").replace("$", "").replace("%", "")
    if not t or t in ("--", "-"):
        return 0.0
    try:
        return float(t)
    except ValueError:
        return 0.0


def _month(v):
    """-> 'YYYY-MM', or '' when the export was pulled without a month segment.

    Google writes 'May 2026'; Meta writes a real date. Both appear.
    """
    if v is None:
        return ""
    if isinstance(v, (dt.datetime, dt.date)):
        return "%04d-%02d" % (v.year, v.month)
    t = _s(v)
    if not t:
        return ""
    m = re.match(r"^(\d{4})-(\d{1,2})", t)
    if m:
        return "%04d-%02d" % (int(m.group(1)), int(m.group(2)))
    m = re.match(r"^([A-Za-z]+)\s+(\d{4})$", t)
    if m and m.group(1).lower() in MONTH_NAMES:
        return "%04d-%02d" % (int(m.group(2)), MONTH_NAMES[m.group(1).lower()])
    return ""


def _iso(v):
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    t = _s(v)
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", t)
    if m:
        return m.group(0)
    # "September 1, 2025"
    m = re.match(r"^([A-Za-z]+)\s+(\d{1,2}),\s*(\d{4})$", t)
    if m and m.group(1).lower() in MONTH_NAMES:
        return "%04d-%02d-%02d" % (int(m.group(3)), MONTH_NAMES[m.group(1).lower()],
                                   int(m.group(2)))
    return ""


def _index(headers, *names):
    """Column index by header name. Returns None when absent."""
    low = [_s(h).lower() for h in headers]
    for want in names:
        w = want.lower()
        for i, h in enumerate(low):
            if h == w:
                return i
    for want in names:                    # loose fallback: substring
        w = want.lower()
        for i, h in enumerate(low):
            if w in h:
                return i
    return None


def _require(idx, label, filename):
    if idx is None:
        raise SpendError(
            "%s: could not find a %r column. Re-export with that column included."
            % (filename, label))
    return idx


def _read_bytes(path_or_stream):
    if hasattr(path_or_stream, "read"):
        return path_or_stream.read()
    with open(path_or_stream, "rb") as fh:
        return fh.read()


# --------------------------------------------------------------------------
# parsers
# --------------------------------------------------------------------------
def parse_google(path_or_stream, filename="google.csv"):
    """Google Ads campaign report -> (rows, meta).

    The export has two preamble lines before the header: a title, then the date
    range in quotes. Zero-cost rows are dropped — the account carries ~40 dormant
    campaigns that would otherwise bulk out the table with nothing in them.
    """
    raw = _read_bytes(path_or_stream)
    for enc in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    else:
        raise SpendError("%s: could not decode the file as text." % filename)

    lines = list(csv.reader(io.StringIO(text)))
    if len(lines) < 3:
        raise SpendError("%s: too short to be a Google Ads export." % filename)

    # Find the header row rather than assuming line 3: report settings change
    # how many preamble lines Google writes.
    head_at = None
    for i, row in enumerate(lines[:8]):
        if _index(row, "Cost") is not None and _index(row, "Campaign") is not None:
            head_at = i
            break
    if head_at is None:
        raise SpendError(
            "%s: no header row with 'Campaign' and 'Cost'. Is this a Google Ads "
            "campaign report?" % filename)

    headers = lines[head_at]
    i_cost = _require(_index(headers, "Cost"), "Cost", filename)
    i_camp = _require(_index(headers, "Campaign"), "Campaign", filename)
    i_type = _index(headers, "Campaign type")
    i_month = _index(headers, "Month")
    i_clicks = _index(headers, "Clicks")
    i_impr = _index(headers, "Impr.", "Impressions")

    period_start = period_end = ""
    for row in lines[:head_at]:
        for cell in row:
            if " - " in _s(cell):
                a, _, b = _s(cell).partition(" - ")
                period_start, period_end = _iso(a), _iso(b)
                break
        if period_start:
            break

    rows, unknown_types = [], set()
    for line in lines[head_at + 1:]:
        if len(line) <= i_cost:
            continue
        cost = _num(line[i_cost])
        if not cost:
            continue
        ctype = _s(line[i_type]) if i_type is not None and i_type < len(line) else ""
        sub = GOOGLE_TYPE_SUB.get(ctype.lower())
        if sub is None:
            sub = GOOGLE_FALLBACK_SUB
            if ctype:
                unknown_types.add(ctype)
        rows.append({
            "platform": GOOGLE,
            "month": _month(line[i_month]) if i_month is not None and i_month < len(line) else "",
            "channel": GOOGLE_CHANNEL,
            "sub_source": sub,
            "campaign": _s(line[i_camp]) if i_camp < len(line) else "",
            "ad_id": "",
            "cost": cost,
            "clicks": int(_num(line[i_clicks])) if i_clicks is not None and i_clicks < len(line) else 0,
            "impressions": int(_num(line[i_impr])) if i_impr is not None and i_impr < len(line) else 0,
        })
    return rows, _meta(rows, filename, GOOGLE, period_start, period_end,
                       unknown_types, "campaign type")


def parse_meta(path_or_stream, filename="meta.xlsx"):
    """Meta Ads raw data report -> (rows, meta). Keeps Ad ID for the coverage check."""
    from openpyxl import load_workbook
    import warnings

    data = _read_bytes(path_or_stream)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        try:
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:                       # noqa: BLE001 - user file
            raise SpendError("%s: could not open as a workbook (%s)."
                             % (filename, exc))
        ws = wb[wb.sheetnames[0]]
        table = list(ws.iter_rows(values_only=True))
        wb.close()

    if len(table) < 2:
        raise SpendError("%s: no data rows." % filename)
    headers = table[0]
    i_cost = _require(_index(headers, "Amount spent (USD)", "Amount spent"),
                      "Amount spent", filename)
    i_plat = _require(_index(headers, "Platform"), "Platform", filename)
    i_camp = _index(headers, "Campaign name")
    i_month = _index(headers, "Month")
    i_ad = _index(headers, "Ad ID")
    i_clicks = _index(headers, "Link clicks")
    i_impr = _index(headers, "Impressions")
    i_from = _index(headers, "Reporting starts")
    i_to = _index(headers, "Reporting ends")

    rows, unknown_plats = [], set()
    period_start = period_end = ""
    for line in table[1:]:
        if len(line) <= i_cost:
            continue
        cost = _num(line[i_cost])
        if not cost:
            continue
        plat = _s(line[i_plat]).lower()
        sub = META_PLATFORM_SUB.get(plat)
        if sub is None:
            sub = META_FALLBACK_SUB
            if plat:
                unknown_plats.add(plat)
        if i_from is not None and not period_start:
            period_start = _iso(line[i_from])
        if i_to is not None:
            period_end = _iso(line[i_to]) or period_end
        rows.append({
            "platform": META,
            "month": _month(line[i_month]) if i_month is not None else "",
            "channel": META_CHANNEL,
            "sub_source": sub,
            "campaign": _s(line[i_camp]) if i_camp is not None else "",
            "ad_id": _s(line[i_ad]) if i_ad is not None else "",
            "cost": cost,
            "clicks": int(_num(line[i_clicks])) if i_clicks is not None else 0,
            "impressions": int(_num(line[i_impr])) if i_impr is not None else 0,
        })
    return rows, _meta(rows, filename, META, period_start, period_end,
                       unknown_plats, "platform")


def _meta(rows, filename, platform, period_start, period_end, unknown, what):
    if not rows:
        raise SpendError("%s: no rows with any spend." % filename)
    warnings = []
    if unknown:
        warnings.append(
            "Unrecognised %s: %s — pooled into the channel's “Other” "
            "sub-source. Tell me and I'll map it properly."
            % (what, ", ".join(sorted(unknown))))
    months = sorted({r["month"] for r in rows if r["month"]})
    if not months:
        warnings.append(
            "No month column, so all spend lands in one bucket for the whole "
            "period. Re-export with a month segment to get cost over time and "
            "cost that follows the date filter.")
    # A resolved sub-source that the taxonomy does not know would silently
    # orphan every dollar in it — the join is by name.
    for r in rows:
        if r["sub_source"] not in dict(taxonomy.TAXONOMY).get(r["channel"], ()):
            raise SpendError(
                "%s: resolved to %r under %r, which is not in taxonomy.py."
                % (filename, r["sub_source"], r["channel"]))
    return {
        "platform": platform, "filename": filename,
        "period_start": period_start or (months[0] + "-01" if months else ""),
        "period_end": period_end,
        "months": months,
        "total_cost": sum(r["cost"] for r in rows),
        "row_count": len(rows),
        "warnings": warnings,
    }


def parse(path_or_stream, platform, filename):
    if platform == GOOGLE:
        return parse_google(path_or_stream, filename)
    if platform == META:
        return parse_meta(path_or_stream, filename)
    raise SpendError("Unknown platform %r." % platform)


# --------------------------------------------------------------------------
# storage
# --------------------------------------------------------------------------
def store(conn, program, rows, meta, sha256=""):
    """Replace this platform's spend for every month the file covers.

    Replace, not merge: a spend export restates a period. Platforms revise
    figures afterwards, so keeping the first number ever seen (the rule that is
    right for funnel stages) would be wrong here.
    """
    now = dt.datetime.now().isoformat(timespec="seconds")
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO spend_uploads (program, platform, filename, sha256,"
        " period_start, period_end, row_count, total_cost, uploaded_at)"
        " VALUES (?,?,?,?,?,?,?,?,?)",
        (program, meta["platform"], meta["filename"], sha256,
         meta["period_start"], meta["period_end"], meta["row_count"],
         meta["total_cost"], now))
    upload_id = cur.lastrowid

    months = sorted({r["month"] for r in rows})
    replaced = 0
    for m in months:
        cur.execute(
            "SELECT COUNT(*) FROM spend WHERE program=? AND platform=? AND month=?",
            (program, meta["platform"], m))
        replaced += cur.fetchone()[0]
        cur.execute(
            "DELETE FROM spend WHERE program=? AND platform=? AND month=?",
            (program, meta["platform"], m))

    for r in rows:
        cur.execute(
            "INSERT INTO spend (program, platform, month, channel, sub_source,"
            " campaign, ad_id, cost, clicks, impressions, upload_id)"
            " VALUES (?,?,?,?,?,?,?,?,?,?,?)"
            " ON CONFLICT(program, platform, month, campaign, ad_id, sub_source)"
            " DO UPDATE SET cost=spend.cost+excluded.cost,"
            "   clicks=spend.clicks+excluded.clicks,"
            "   impressions=spend.impressions+excluded.impressions,"
            "   upload_id=excluded.upload_id",
            (program, r["platform"], r["month"], r["channel"], r["sub_source"],
             r["campaign"], r["ad_id"], r["cost"], r["clicks"], r["impressions"],
             upload_id))
    conn.commit()
    out = dict(meta)
    out.update({"upload_id": upload_id, "replaced": replaced,
                "program": program, "stored": len(rows)})
    return out


def meta_ad_coverage(conn, program):
    """How much Meta spend is on ads the funnel actually saw.

    Meta tags its links `utm_content={{ad.id}}`, so utm_content joins straight to
    the export's Ad ID. Nothing in the rollup depends on this — it is here as
    evidence that the spend file and the ping log describe the same traffic. A
    sharp drop means they have diverged.
    """
    spend_rows = conn.execute(
        "SELECT ad_id, SUM(cost) c FROM spend"
        " WHERE program=? AND platform=? AND ad_id <> '' GROUP BY ad_id",
        (program, META)).fetchall()
    if not spend_rows:
        return None
    by_ad = {r["ad_id"]: r["c"] for r in spend_rows}
    ping_rows = conn.execute(
        "SELECT p.content ad, COUNT(*) n FROM pings p"
        " JOIN applicants a ON a.id = p.applicant_id"
        " WHERE a.program=? AND p.channel=? GROUP BY p.content",
        (program, META_CHANNEL)).fetchall()
    pings = {r["ad"]: r["n"] for r in ping_rows}
    matched = set(by_ad) & set(pings)
    total_cost = sum(by_ad.values())
    total_pings = sum(pings.values()) or 1
    return {
        "ads": len(by_ad),
        "matched_ads": len(matched),
        "spend_total": total_cost,
        "spend_matched": sum(by_ad[a] for a in matched),
        "spend_pct": (sum(by_ad[a] for a in matched) / total_cost) if total_cost else 0.0,
        "pings_total": sum(pings.values()),
        "pings_matched": sum(pings[a] for a in matched),
        "pings_pct": sum(pings[a] for a in matched) / total_pings,
        "unmatched_spend": sum(v for k, v in by_ad.items() if k not in pings),
    }
