"""Regression tests wiring the analysis_engine.py docstring numbers into CI.

Per CLAUDE.md the app is only correct when it reproduces these figures against
the real files in sample_data/. Three layers are checked:

  1. the canonical engine still produces the documented numbers;
  2. the vendored copy of the engine has not drifted from the handoff original;
  3. the DB round-trip (ingest -> SQLite -> metrics.build_matrix) reproduces the
     canonical numbers exactly, and is idempotent under re-upload.
"""
import hashlib
from collections import defaultdict
import os
import sqlite3

import openpyxl
import pytest

from app import db, filters, ingest, metrics, programs, spend, taxonomy
from app.analysis_engine import build_matrix as engine_build_matrix
from app.analysis_engine import ft_stages, summer_stages
from conftest import CANONICAL_ENGINE, FT_FILE, SUMMER_FILE

FT_STAGES = ["started", "submitted", "aud_req", "aud_comp", "admitted"]

# ---- documented expectations (analysis_engine.py docstring) -----------------
FT_TOTALS = {
    "started": 8436, "submitted": 1240, "aud_req": 875,
    "aud_comp": 460, "admitted": 350,
}
# (channel, sub) -> (n, admitted)
FT_CHANNELS = {
    ("Google (Paid)", None): (3520, 229),
    ("Google (Paid)", "PMax"): (3279, 226),
    ("Google (Paid)", "Search"): (350, 16),
    ("Google (Organic)", None): (1634, 264),
    ("Meta Paid Social (IG/FB)", None): (2757, 28),
    ("Meta Paid Social (IG/FB)", "Instagram"): (1811, 21),
    ("Meta Paid Social (IG/FB)", "Facebook"): (943, 6),
    ("Meta Paid Social (IG/FB)", "Untagged (broken tag)"): (20, 1),
    ("Organic/Other Search", None): (205, 45),
    ("Spotify (Paid Audio)", None): (1, 0),
    ("Partner/Referral", None): (2, None),
    ("Unresolved/Other", None): (3, None),
}
SUMMER_TOTALS = {"started": 4483, "submitted": 954, "accepted": 387}


def _rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = list(wb["Export"].iter_rows(values_only=True))[1:]
    wb.close()
    return out


# --------------------------------------------------------------------------
# 1. canonical engine
# --------------------------------------------------------------------------
def test_canonical_engine_reproduces_documented_ft_numbers():
    res = engine_build_matrix(_rows(FT_FILE), (15, 16, 17, 18), ft_stages, FT_STAGES)
    assert _canonical(res["totals"]) == FT_TOTALS
    for key, (n, admits) in FT_CHANNELS.items():
        assert key in res["matrix"], "missing channel row %r" % (key,)
        assert res["matrix"][key]["n"] == n, "n mismatch for %r" % (key,)
        if admits is not None:
            assert res["matrix"][key]["counts"]["admitted"] == admits, \
                "admits mismatch for %r" % (key,)


def test_documented_subsource_admit_overlap():
    """Docstring notes PMax+Search admits exceed the parent by 13 (any-touch)."""
    res = engine_build_matrix(_rows(FT_FILE), (15, 16, 17, 18), ft_stages, FT_STAGES)
    m = res["matrix"]
    pmax = m[("Google (Paid)", "PMax")]["counts"]["admitted"]
    search = m[("Google (Paid)", "Search")]["counts"]["admitted"]
    parent = m[("Google (Paid)", None)]["counts"]["admitted"]
    assert pmax + search - parent == 13


def test_other_paid_subsource_is_present_once():
    """The docstring comment says 'observed count: 0' for Google (Paid)/Other
    paid, but the sample file contains exactly one such ping
    (source=Google, medium=paid, no campaign). The rule is right; the comment
    is stale. Pinned here so the real number is the one under test."""
    res = engine_build_matrix(_rows(FT_FILE), (15, 16, 17, 18), ft_stages, FT_STAGES)
    assert res["matrix"][("Google (Paid)", "Other paid")]["n"] == 1


def test_canonical_engine_reproduces_summer_totals():
    res = engine_build_matrix(_rows(SUMMER_FILE), (10, 12, 13, 14), summer_stages,
                              ["started", "submitted", "accepted"])
    assert res["totals"] == SUMMER_TOTALS


def test_no_utm_row_is_present_and_material():
    """~13% of applicants have no UTM at all; the row must never be dropped."""
    res = engine_build_matrix(_rows(FT_FILE), (15, 16, 17, 18), ft_stages, FT_STAGES)
    no_utm = res["matrix"][(taxonomy.NO_UTM, None)]["n"]
    assert no_utm == 1133
    share = no_utm / FT_TOTALS["started"]
    assert 0.12 < share < 0.15, "No-UTM share drifted from the documented ~13%%"


# --------------------------------------------------------------------------
# 2. vendored copy has not drifted
# --------------------------------------------------------------------------
def test_vendored_engine_matches_handoff_original():
    with open(CANONICAL_ENGINE, "rb") as fh:
        canonical = hashlib.sha256(fh.read()).hexdigest()
    vendored_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                 "..", "app", "analysis_engine.py")
    with open(vendored_path, "rb") as fh:
        vendored = hashlib.sha256(fh.read()).hexdigest()
    assert vendored == canonical, (
        "app/analysis_engine.py has drifted from the handoff reference copy. "
        "Re-copy it rather than editing the vendored file."
    )


# --------------------------------------------------------------------------
# 3. DB round-trip
# --------------------------------------------------------------------------
@pytest.fixture(scope="module")
def ft_db():
    conn = db.connect(":memory:")
    with open(FT_FILE, "rb") as fh:
        payload = fh.read()
    ingest.ingest(conn, FT_FILE, "ft", os.path.basename(FT_FILE),
                  ingest.sha256_of(payload))
    yield conn
    conn.close()



def _canonical(totals):
    """Stage totals restricted to the stages the vendored engine owns.

    FT_TOTALS is the handoff's published numbers and must never move. `enrolled`
    is derived in programs.py, not by the engine, so it is compared on its own
    (see test_enrolled_stage_nests_inside_admitted) rather than by loosening
    this guard.
    """
    return {k: v for k, v in totals.items() if k in FT_TOTALS}


def _db_matrix(conn, program_key, touch="any"):
    program = programs.get(program_key)
    apps, flags = metrics.load_population(conn, program, "program = ?", [program_key])
    pings = metrics.load_pings(conn, [a["id"] for a in apps])
    return metrics.build_matrix(program, apps, flags, pings, touch=touch)


def test_db_roundtrip_reproduces_canonical_ft_numbers(ft_db):
    res = _db_matrix(ft_db, "ft")
    assert _canonical(res["totals"]) == FT_TOTALS
    for key, (n, admits) in FT_CHANNELS.items():
        row = res["by_key"].get(key)
        assert row is not None, "missing channel row %r after DB round-trip" % (key,)
        assert row["n"] == n, "n mismatch for %r" % (key,)
        if admits is not None:
            assert row["counts"]["admitted"] == admits, "admits mismatch for %r" % (key,)


def test_db_matrix_matches_engine_matrix_row_for_row(ft_db):
    """Strongest form: every channel row agrees with the canonical engine."""
    engine = engine_build_matrix(_rows(FT_FILE), (15, 16, 17, 18), ft_stages, FT_STAGES)
    dbres = _db_matrix(ft_db, "ft")
    assert set(engine["matrix"].keys()) == set(dbres["by_key"].keys())
    for key, want in engine["matrix"].items():
        got = dbres["by_key"][key]
        assert got["n"] == want["n"], "n mismatch %r" % (key,)
        assert _canonical(got["counts"]) == want["counts"], \
            "counts mismatch %r" % (key,)


def test_taxonomy_declares_every_channel_the_engine_emits(ft_db):
    """A new classification rule must also get a display slot."""
    assert _db_matrix(ft_db, "ft")["undeclared_keys"] == []


def test_reupload_is_idempotent():
    """Re-ingesting an overlapping export must not double-count anything."""
    conn = db.connect(":memory:")
    with open(FT_FILE, "rb") as fh:
        digest = ingest.sha256_of(fh.read())

    first = ingest.ingest(conn, FT_FILE, "ft", "first.xlsx", digest)
    before = _db_matrix(conn, "ft")
    counts_before = (
        conn.execute("SELECT COUNT(*) FROM applicants").fetchone()[0],
        conn.execute("SELECT COUNT(*) FROM pings").fetchone()[0],
    )

    second = ingest.ingest(conn, FT_FILE, "ft", "second.xlsx", digest)
    after = _db_matrix(conn, "ft")
    counts_after = (
        conn.execute("SELECT COUNT(*) FROM applicants").fetchone()[0],
        conn.execute("SELECT COUNT(*) FROM pings").fetchone()[0],
    )

    assert first["applicants_new"] == 8436
    assert second["applicants_new"] == 0
    assert second["applicants_updated"] == 8436
    assert second["pings_new"] == 0
    assert second["pings_duplicate"] == first["pings_new"]
    assert counts_before == counts_after
    assert after["totals"] == before["totals"]
    for key, row in before["by_key"].items():
        assert after["by_key"][key]["n"] == row["n"]
        assert after["by_key"][key]["counts"] == row["counts"]
    conn.close()


def test_summer_db_roundtrip():
    conn = db.connect(":memory:")
    ingest.ingest(conn, SUMMER_FILE, "summer", os.path.basename(SUMMER_FILE), "")
    res = _db_matrix(conn, "summer")
    assert res["totals"] == SUMMER_TOTALS
    conn.close()


def test_programs_stay_separated():
    """Full-Time and Summer must never be merged."""
    conn = db.connect(":memory:")
    ingest.ingest(conn, FT_FILE, "ft", "ft.xlsx", "")
    ingest.ingest(conn, SUMMER_FILE, "summer", "summer.xlsx", "")
    ft = _db_matrix(conn, "ft")
    su = _db_matrix(conn, "summer")
    assert _canonical(ft["totals"]) == FT_TOTALS
    assert su["totals"] == SUMMER_TOTALS
    total_rows = conn.execute("SELECT COUNT(*) FROM applicants").fetchone()[0]
    assert total_rows == 8436 + 4483
    conn.close()


# --------------------------------------------------------------------------
# derived views
# --------------------------------------------------------------------------
def test_overall_funnel_matches_stage_totals(ft_db):
    program = programs.get("ft")
    apps, flags = metrics.load_population(ft_db, program, "program = ?", ["ft"])
    overall = metrics.overall_funnel(program, flags)
    assert overall["population"] == 8436
    assert _canonical(overall["counts"]) == FT_TOTALS
    assert [s["n"] for s in overall["steps"]] == [8436, 1240, 875, 460, 350, 16]
    # step conversions are relative to the previous stage
    submitted = overall["steps"][1]
    assert abs(submitted["pct_of_prev"] - 1240 / 8436) < 1e-12


def test_first_touch_rows_sum_to_stage_total(ft_db):
    """Unlike any-touch, first-touch assigns each person exactly one parent."""
    res = _db_matrix(ft_db, "ft", touch="first")
    parents = [r for r in res["rows"] if r["is_parent"]]
    assert sum(r["n"] for r in parents) == 8436
    assert sum(r["counts"]["admitted"] for r in parents) == FT_TOTALS["admitted"]


def test_any_touch_rows_overlap_and_exceed_stage_total(ft_db):
    """Guards the golden rule: any-touch rows must NOT be summable."""
    res = _db_matrix(ft_db, "ft", touch="any")
    parents = [r for r in res["rows"] if r["is_parent"]]
    assert sum(r["n"] for r in parents) > 8436


def test_top_utm_breakdown_respects_stage_totals(ft_db):
    program = programs.get("ft")
    apps, flags = metrics.load_population(ft_db, program, "program = ?", ["ft"])
    pings = metrics.load_pings(ft_db, [a["id"] for a in apps])
    for field in ("campaign", "content", "source", "medium"):
        out = metrics.top_utm_breakdown(program, apps, flags, pings, field, limit=10)
        assert _canonical(out["totals"]) == FT_TOTALS
        assert len(out["rows"]) <= 10
        for row in out["rows"]:
            assert row["counts"]["admitted"] <= FT_TOTALS["admitted"]
            assert row["n"] >= row["counts"]["admitted"]


def test_unknown_utms_are_recorded_for_review():
    """The brief asks to be told when a UTM doesn't fit a category."""
    conn = db.connect(":memory:")
    res = ingest.ingest(conn, FT_FILE, "ft", "ft.xlsx", "")
    rows = conn.execute(
        "SELECT source, medium, campaign, ping_count FROM unknown_utms"
    ).fetchall()
    assert rows, "expected the unclassified pings to be flagged"
    combos = {(r["source"], r["medium"]) for r in rows}
    assert ("Chrome", "Extension:omnibox") in combos
    assert sum(r["ping_count"] for r in rows) == 4
    assert res["new_unknown_utms"]
    conn.close()


def test_unknown_utm_counts_do_not_inflate_on_reupload():
    """The review queue is recomputed from deduplicated pings, so uploading the
    same export twice must not make an unclassified combo look more common."""
    conn = db.connect(":memory:")
    ingest.ingest(conn, FT_FILE, "ft", "first.xlsx", "")
    first = conn.execute(
        "SELECT source, medium, campaign, ping_count, applicant_count"
        " FROM unknown_utms ORDER BY source, medium, campaign"
    ).fetchall()
    ingest.ingest(conn, FT_FILE, "ft", "second.xlsx", "")
    second = conn.execute(
        "SELECT source, medium, campaign, ping_count, applicant_count"
        " FROM unknown_utms ORDER BY source, medium, campaign"
    ).fetchall()
    assert [tuple(r) for r in first] == [tuple(r) for r in second]
    assert sum(r["ping_count"] for r in second) == 4
    conn.close()


def test_acknowledged_unknown_utm_survives_a_reupload():
    """Marking a combo reviewed must not be undone by re-ingesting old data."""
    conn = db.connect(":memory:")
    ingest.ingest(conn, FT_FILE, "ft", "first.xlsx", "")
    conn.execute("UPDATE unknown_utms SET acknowledged=1")
    conn.commit()
    ingest.ingest(conn, FT_FILE, "ft", "second.xlsx", "")
    still = conn.execute(
        "SELECT COUNT(*) FROM unknown_utms WHERE acknowledged=0"
    ).fetchone()[0]
    assert still == 0
    conn.close()


def test_shifted_columns_are_rejected_not_silently_misread():
    """A column inserted mid-export must fail loudly.

    Everything is parsed by index, so a positional shift would read dates as
    decisions and UTMs as postcodes with no error at all. A BFA program is
    expected to add a column, so this guard matters.
    """
    import io
    import openpyxl as _ox

    src = _ox.load_workbook(FT_FILE, read_only=True, data_only=True)
    rows = []
    for i, row in enumerate(src["Export"].iter_rows(values_only=True)):
        rows.append(list(row))
        if i > 40:
            break
    src.close()

    # Insert a new column in the MIDDLE, as a Slate layout change might.
    out = _ox.Workbook()
    ws = out.active
    ws.title = "Export"
    for i, row in enumerate(rows):
        row = list(row)
        row.insert(5, "BFA 4-Year" if i == 0 else "Y")
        ws.append(row)
    buf = io.BytesIO()
    out.save(buf)
    buf.seek(0)

    conn = db.connect(":memory:")
    with pytest.raises(ingest.IngestError) as err:
        ingest.ingest(conn, buf, "ft", "shifted.xlsx", "")
    assert "match none of the" in str(err.value)
    assert conn.execute("SELECT COUNT(*) FROM applicants").fetchone()[0] == 0
    conn.close()


def test_appended_column_is_accepted_with_a_warning():
    """A column added at the END is harmless and must not block an import."""
    import io
    import openpyxl as _ox

    src = _ox.load_workbook(FT_FILE, read_only=True, data_only=True)
    rows = [list(r) for i, r in
            enumerate(src["Export"].iter_rows(values_only=True)) if i <= 40]
    src.close()

    out = _ox.Workbook()
    ws = out.active
    ws.title = "Export"
    for i, row in enumerate(rows):
        ws.append(list(row) + ["BFA Track" if i == 0 else "4-Year"])
    buf = io.BytesIO()
    out.save(buf)
    buf.seek(0)

    conn = db.connect(":memory:")
    res = ingest.ingest(conn, buf, "ft", "appended.xlsx", "")
    assert res["applicants_new"] == 40
    assert any("beyond the" in w for w in res["warnings"])
    conn.close()


# --------------------------------------------------------------------------
# native referrer ping log (the third export)
# --------------------------------------------------------------------------
REF_HEADERS = [
    "Person Reference ID", "Ping Referrer", "Ping Duration (seconds)",
    "Ping Timestamp", "Ping URL", "Ping UTM Campaign", "Ping UTM Medium",
    "Ping UTM Content", "Ping UTM Source", "Ping UTM Term",
]


def _ref_workbook(rows):
    import io
    import openpyxl as _ox
    wb = _ox.Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append(REF_HEADERS)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_ref_domain_extraction():
    assert ingest.ref_domain("https://www.google.com/search?q=x") == "google.com"
    assert ingest.ref_domain("https://apply.aada.edu/apply/") == "apply.aada.edu"
    # Android app referrers arrive bare, not as URLs — Gmail opens depend on this
    assert ingest.ref_domain("com.google.android.gm") == "com.google.android.gm"
    assert ingest.ref_domain("") == ""
    assert ingest.ref_domain(None) == ""
    assert ingest.is_internal("apply.aada.edu")
    assert ingest.is_internal("aada.edu")
    assert not ingest.is_internal("google.com")
    assert not ingest.is_internal("notaada.edu.evil.com")


def test_referral_ingest_classifies_and_dedupes():
    rows = [
        ["001", "https://www.google.com/", 5, "2026-01-02 10:00:00",
         "https://apply.aada.edu/apply/", "c1", "cpc", "ct", "google", "kw1"],
        ["001", "com.google.android.gm", 3, "2026-01-02 10:05:00",
         "https://apply.aada.edu/apply/", None, None, None, None, None],
        ["002", "https://apply.aada.edu/account/login", 9, "2026-01-03 11:00:00",
         "https://apply.aada.edu/apply/", None, None, None, None, None],
    ]
    conn = db.connect(":memory:")
    res = ingest.ingest_referrals(conn, _ref_workbook(rows), "ref.xlsx", "abc")
    assert res["pings_new"] == 3
    assert res["people"] == 2

    got = {r["domain"]: (r["internal"], r["has_utm"]) for r in
           conn.execute("SELECT domain, internal, has_utm FROM ref_pings")}
    assert got["google.com"] == (0, 1)
    assert got["com.google.android.gm"] == (0, 0)   # untagged external -> the win
    assert got["apply.aada.edu"] == (1, 0)          # our own site, not a source

    # re-upload is idempotent
    again = ingest.ingest_referrals(conn, _ref_workbook(rows), "ref2.xlsx", "abc")
    assert again["pings_new"] == 0
    assert again["pings_duplicate"] == 3
    assert conn.execute("SELECT COUNT(*) FROM ref_pings").fetchone()[0] == 3
    assert any("already uploaded" in w for w in again["warnings"])
    conn.close()


def test_referral_ingest_rejects_the_wrong_export():
    """Uploading a Ping Data file as a referrer log must fail, not half-import."""
    conn = db.connect(":memory:")
    with pytest.raises(ingest.IngestError) as err:
        ingest.ingest_referrals(conn, FT_FILE, "wrong.xlsx", "")
    assert "referrer ping log" in str(err.value)
    assert conn.execute("SELECT COUNT(*) FROM ref_pings").fetchone()[0] == 0
    conn.close()


def test_referrer_report_joins_to_the_funnel_without_touching_it():
    """The ping log must not alter any funnel figure, and must still be able to
    report stage outcomes per referrer."""
    conn = db.connect(":memory:")
    ingest.ingest(conn, FT_FILE, "ft", "ft.xlsx", "")
    before = _db_matrix(conn, "ft")

    gid = conn.execute(
        "SELECT global_id FROM applicants WHERE program='ft' AND st_admitted=1 LIMIT 1"
    ).fetchone()[0]
    ingest.ingest_referrals(conn, _ref_workbook([
        [gid, "https://chatgpt.com/", 4, "2026-02-01 09:00:00",
         "https://apply.aada.edu/apply/", None, None, None, None, None],
    ]), "ref.xlsx", "")

    after = _db_matrix(conn, "ft")
    assert _canonical(after["totals"]) == _canonical(before["totals"]) == FT_TOTALS
    for key, row in before["by_key"].items():
        assert after["by_key"][key]["counts"] == row["counts"]

    program = programs.get("ft")
    rep = metrics.top_referrers(conn, program, "program = ?", ["ft"],
                                group="domain", scope="untagged")
    hit = [r for r in rep["rows"] if r["label"] == "chatgpt.com"]
    assert hit, "expected the untagged external referrer to be reported"
    assert hit[0]["counts"]["admitted"] == 1
    assert hit[0]["matched"] >= 1

    summary = metrics.referrer_summary(conn, program, "program = ?", ["ft"])
    assert summary["untagged_external"] == 1
    conn.close()


# --------------------------------------------------------------------------
# tag timeline
# --------------------------------------------------------------------------
def _tl_parents(conn, program):
    """The parent-only selection the timeline opens with."""
    tree = metrics.timeline_tree(metrics.timeline_facets(conn, program))
    return [p["name"] for p in tree]


def test_fiscal_year_boundaries():
    """Fiscal year runs 1 Sept -> 31 Aug, so Jan-Aug belong to the year that
    started the previous September. August must land INSIDE a year, not between
    two of them."""
    assert metrics.fiscal_year_of("2025-09-01") == 2025
    assert metrics.fiscal_year_of("2025-12-31") == 2025
    assert metrics.fiscal_year_of("2026-01-01") == 2025
    assert metrics.fiscal_year_of("2026-07-31") == 2025
    assert metrics.fiscal_year_of("2026-08-31") == 2025   # the whole point
    assert metrics.fiscal_year_of("2026-09-01") == 2026
    assert metrics.fiscal_year_label(2025) == "FY 2025/26"


def test_timeline_monthly_buckets_match_raw_counts(ft_db):
    """Bucketing must be exact — every tag in exactly one bucket, none lost.

    Checked two ways, because only 8 categories are DRAWN (palette limit) while
    every category is still COUNTED:
      * grand_total covers all categories and must equal the raw total;
      * one category's per-bucket series must match raw SQL for that category.
    """
    program = programs.get("ft")
    tl = metrics.tag_timeline(ft_db, program, "program = ?", ["ft"], bucket="month",
                              picked=_tl_parents(ft_db, program))
    assert tl["labels"] == ["Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar",
                            "Apr", "May", "Jun", "Jul", "Aug"]

    raw_total = ft_db.execute(
        "SELECT COUNT(*) FROM pings p JOIN applicants a ON a.id = p.applicant_id"
        " WHERE a.program='ft'"
        "   AND COALESCE(NULLIF(p.ts,''), a.started_date) <> ''").fetchone()[0]
    assert tl["grand_total"] == raw_total, "every tag must be counted"

    raw = dict(ft_db.execute(
        "SELECT (CAST(strftime('%m', d) AS INTEGER) - 9 + 12) % 12 AS b, COUNT(*)"
        " FROM (SELECT substr(COALESCE(NULLIF(p.ts,''), a.started_date),1,10) d"
        "         FROM pings p JOIN applicants a ON a.id = p.applicant_id"
        "        WHERE a.program='ft' AND p.channel = 'Google (Paid)'"
        "          AND COALESCE(NULLIF(p.ts,''), a.started_date) <> '')"
        " GROUP BY b").fetchall())
    gp = [s for s in tl["series"] if s["group"] == "Google (Paid)"]
    agg = [0] * 12
    for s in gp:
        for i, v in enumerate(s["data"]):
            agg[i] += v
    assert agg == [raw.get(i, 0) for i in range(12)]

    aug = ft_db.execute(
        "SELECT COUNT(*) FROM pings p JOIN applicants a ON a.id = p.applicant_id"
        " WHERE a.program='ft'"
        "   AND substr(COALESCE(NULLIF(p.ts,''), a.started_date),6,2) = '08'"
    ).fetchone()[0]
    assert aug > 0
    drawn_aug = sum(s["data"][11] for s in tl["series"])
    assert drawn_aug > 0, "August tags must be inside the fiscal year, not stranded"


def test_timeline_dates_the_tag_and_falls_back_to_app_start(ft_db):
    """The fallback must be counted, and must cover every undated tag."""
    program = programs.get("ft")
    tl = metrics.tag_timeline(ft_db, program, "program = ?", ["ft"],
                              bucket="month", max_series=99,
                              picked=_tl_parents(ft_db, program))
    undated = ft_db.execute(
        "SELECT COUNT(*) FROM pings p JOIN applicants a ON a.id = p.applicant_id"
        " WHERE a.program='ft' AND p.ts = ''").fetchone()[0]
    assert tl["approx"] == undated
    # every undated ping still has a start date, so nothing is lost
    assert ft_db.execute(
        "SELECT COUNT(*) FROM pings p JOIN applicants a ON a.id = p.applicant_id"
        " WHERE a.program='ft' AND p.ts = '' AND a.started_date = ''"
    ).fetchone()[0] == 0


def test_timeline_marks_only_the_newest_year_current(ft_db):
    program = programs.get("ft")
    tl = metrics.tag_timeline(ft_db, program, "program = ?", ["ft"], bucket="month")
    assert tl["years"] == sorted(tl["years"], reverse=True)
    for s in tl["series"]:
        assert s["current"] == (s["fy"] == tl["newest"])


def test_timeline_respects_category_and_bucket_choices(ft_db):
    """One picker now covers parents AND sub-sources; there is no mode toggle."""
    program = programs.get("ft")
    parents = _tl_parents(ft_db, program)

    only = metrics.tag_timeline(ft_db, program, "program = ?", ["ft"],
                                picked=["Google (Paid)"], bucket="month")
    assert {s["group"] for s in only["series"]} == {"Google (Paid)"}

    # a sub-source on its own, addressed by its composed name
    pmax = taxonomy.sub_name("Google (Paid)", "PMax")
    sub_only = metrics.tag_timeline(ft_db, program, "program = ?", ["ft"],
                                    picked=[pmax], bucket="month")
    assert {s["group"] for s in sub_only["series"]} == {pmax}

    # parent and child together: the parent covers the WHOLE channel, so it
    # overlaps its own child rather than excluding it
    both = metrics.tag_timeline(ft_db, program, "program = ?", ["ft"],
                                picked=["Google (Paid)", pmax], bucket="month")
    # one series PER (group, fiscal year), so totals must be summed per group
    got = defaultdict(int)
    for sv in both["series"]:
        got[sv["group"]] += sv["total"]
    assert set(got) == {"Google (Paid)", pmax}
    assert got["Google (Paid)"] > got[pmax] > 0

    # and the parent total is unchanged by the child also being selected
    assert got["Google (Paid)"] == sum(sv["total"] for sv in only["series"])

    for bucket, n in (("month", 12), ("week", 53), ("day", 366)):
        tl = metrics.tag_timeline(ft_db, program, "program = ?", ["ft"],
                                  picked=parents, bucket=bucket)
        assert len(tl["labels"]) == n
        assert all(len(s["data"]) == n for s in tl["series"])

    none_sel = metrics.tag_timeline(ft_db, program, "program = ?", ["ft"],
                                    bucket="month", select_none=True)
    assert none_sel["series"] == [] and none_sel["empty"] is True


def test_timeline_headline_ignores_the_parent_sub_overlap(ft_db):
    """`ent` lists a ping twice when it has a sub-source (once under the parent,
    once under the sub). The headline must come from the un-duplicated rows or
    it would roughly double."""
    program = programs.get("ft")
    raw = ft_db.execute(
        "SELECT COUNT(*) FROM pings p JOIN applicants a ON a.id = p.applicant_id"
        " WHERE a.program='ft'"
        "   AND COALESCE(NULLIF(p.ts,''), a.started_date) <> ''").fetchone()[0]
    for picked in ([p["name"] for p in metrics.timeline_tree(
                        metrics.timeline_facets(ft_db, program))],
                   ["Google (Paid)", taxonomy.sub_name("Google (Paid)", "PMax")]):
        tl = metrics.tag_timeline(ft_db, program, "program = ?", ["ft"],
                                  picked=picked, bucket="month")
        assert tl["tags_total"] == raw


def test_started_apps_reference_band(ft_db):
    """The grey reference band: applications started, same buckets as the tags,
    dated on the application's own start date."""
    program = programs.get("ft")
    for bucket, n in (("month", 12), ("week", 53), ("day", 366)):
        band = metrics.started_apps_series(ft_db, program, "program = ?", ["ft"],
                                           bucket=bucket)
        assert band["series"], "expected a band for %s" % bucket
        assert all(len(s["data"]) == n for s in band["series"])

    band = metrics.started_apps_series(ft_db, program, "program = ?", ["ft"],
                                       bucket="month")
    dated = ft_db.execute(
        "SELECT COUNT(*) FROM applicants WHERE program='ft' AND started_date <> ''"
    ).fetchone()[0]
    assert band["total"] == dated, "every dated application must land in a bucket"

    # cross-check one bucket against raw SQL
    raw = dict(ft_db.execute(
        "SELECT (CAST(strftime('%m', d) AS INTEGER) - 9 + 12) % 12 AS b, COUNT(*)"
        " FROM (SELECT substr(started_date,1,10) d FROM applicants"
        "        WHERE program='ft' AND started_date <> '')"
        " GROUP BY b").fetchall())
    agg = [0] * 12
    for s in band["series"]:
        for i, v in enumerate(s["data"]):
            agg[i] += v
    assert agg == [raw.get(i, 0) for i in range(12)]

    # it must honour the filter bar like everything else
    flt = filters.Filters(program, terms=["Fall 2026"])
    scoped = metrics.started_apps_series(ft_db, program, flt.where, flt.params,
                                         bucket="month")
    assert 0 < scoped["total"] < band["total"]

    # and the ratio the UI quotes is real: tags far outrun starts
    tags = metrics.tag_timeline(ft_db, program, "program = ?", ["ft"], bucket="month")
    assert tags["tags_total"] > band["total"] * 3


def test_timeline_never_reuses_a_colour_for_two_channels(ft_db):
    """Two different PARENT channels must never share a hue — a repeated colour
    reads as the same channel. A sub-source deliberately DOES share its parent's
    slot (drawn as a desaturated, dotted variant), because it belongs to that
    parent; that is the one intended kind of sharing."""
    program = programs.get("ft")
    tl = metrics.tag_timeline(ft_db, program, "program = ?", ["ft"], bucket="month")
    by_colour = {}
    for s in tl["series"]:
        assert s["colour_index"] is None or 0 <= s["colour_index"] < 8
        # a sub-source shares its parent's slot on purpose
        parent = s["group"].split(taxonomy.SUB_SEP)[0]
        by_colour.setdefault(s["colour_index"], set()).add(parent)
    for idx, parents in by_colour.items():
        if idx is None:
            continue          # the grey tail is shared by design
        assert len(parents) == 1, "colour %d shared by %s" % (idx, sorted(parents))
    assert len({s["group"] for s in tl["series"]}) <= 8
    # asking for more than the palette holds is clamped, not silently cycled
    wide = metrics.tag_timeline(ft_db, program, "program = ?", ["ft"],
                                bucket="month", max_series=99)
    assert len({s["group"] for s in wide["series"]}) <= 8
    assert wide["dropped"], "extra categories must be reported, not dropped silently"


def test_channel_makeup_keeps_no_utm_and_overlaps(ft_db):
    """Makeup is stage penetration: rows overlap past 100%, and the untracked
    row is never dropped."""
    program = programs.get("ft")
    apps, flags = metrics.load_population(ft_db, program, "program = ?", ["ft"])
    pings = metrics.load_pings(ft_db, [a["id"] for a in apps])
    matrix = metrics.build_matrix(program, apps, flags, pings, "any")
    mk = metrics.channel_makeup(program, matrix, stage="admitted")

    assert mk["final_total"] == FT_TOTALS["admitted"]
    labels = [r["channel"] for r in mk["rows"]]
    assert taxonomy.NO_UTM in labels, "the untracked row must survive"
    # ranked by share, and any-touch so the shares exceed 100%
    shares = [r["share"] for r in mk["rows"]]
    assert shares == sorted(shares, reverse=True)
    assert mk["sum_share"] > 1.0
    top = mk["rows"][0]
    assert top["channel"] == "Google (Organic)"
    assert top["final_n"] == 264
    assert abs(top["share"] - 264 / 350) < 1e-12


# --------------------------------------------------------------------------
# filtering on a MISSING value
# --------------------------------------------------------------------------
def _count(conn, flt):
    return conn.execute(
        "SELECT COUNT(*) FROM applicants WHERE " + flt.where, flt.params).fetchone()[0]


def test_blank_values_are_filterable(ft_db):
    """A fifth to a half of rows have no term / region / programme. Before the
    sentinel they were unreachable — the facet lists only offered non-empty
    values, so filtering any of those dimensions silently dropped them."""
    program = programs.get("ft")
    for field, attr in (("term", "terms"), ("region", "regions"),
                        ("emphasis", "emphases")):
        blank = ft_db.execute(
            "SELECT COUNT(*) FROM applicants WHERE program='ft' AND %s=''" % field
        ).fetchone()[0]
        assert blank > 0, "expected blanks in %s" % field
        flt = filters.Filters(program, **{attr: [filters.NONE_TOKEN]})
        assert _count(ft_db, flt) == blank


def test_blank_and_real_values_combine_as_a_union(ft_db):
    program = programs.get("ft")
    blank = _count(ft_db, filters.Filters(program, terms=[filters.NONE_TOKEN]))
    fall = _count(ft_db, filters.Filters(program, terms=["Fall 2026"]))
    both = _count(ft_db, filters.Filters(
        program, terms=[filters.NONE_TOKEN, "Fall 2026"]))
    assert both == blank + fall


def test_untracked_channel_is_filterable(ft_db):
    """"(no UTM)" is not a stored channel — those applicants have no ping rows
    at all — so it must resolve to NOT EXISTS, not an IN."""
    program = programs.get("ft")
    flt = filters.Filters(program, channels=[filters.NONE_TOKEN])
    expected = ft_db.execute(
        "SELECT COUNT(*) FROM applicants a WHERE a.program='ft'"
        " AND NOT EXISTS (SELECT 1 FROM pings p WHERE p.applicant_id = a.id)"
    ).fetchone()[0]
    assert expected > 0
    assert _count(ft_db, flt) == expected
    # and it unions with a real channel rather than replacing it
    gp = _count(ft_db, filters.Filters(program, channels=["Google (Paid)"]))
    both = _count(ft_db, filters.Filters(
        program, channels=["Google (Paid)", filters.NONE_TOKEN]))
    assert both == gp + expected


def test_term_facet_is_alphabetical_with_blank_last(ft_db):
    program = programs.get("ft")
    facets = filters.facet_values(ft_db, program)
    values = [v for v, _ in facets["terms"]]
    assert values[-1] == filters.NONE_TOKEN, "the catch-all belongs last"
    real = values[:-1]
    assert real == sorted(real, key=lambda s: s.lower())
    # high-cardinality columns stay volume-ranked, where the big ones matter
    counts = [n for v, n in facets["regions"] if v != filters.NONE_TOKEN]
    assert counts == sorted(counts, reverse=True)


def test_sentinel_never_reaches_the_ui(ft_db):
    """`__none__` is a wire value; the bar and picker must show a real label."""
    program = programs.get("ft")
    facets = filters.facet_values(ft_db, program)
    flt = filters.Filters(program, terms=[filters.NONE_TOKEN],
                          channels=[filters.NONE_TOKEN])
    blob = repr(flt.tokens()) + flt.summary()
    assert filters.NONE_TOKEN not in blob
    assert "(no term)" in blob and "untracked" in blob

    spec = filters.describe(program, facets, flt)
    for dim in spec["dimensions"]:
        for v in dim.get("values", []):
            if v["v"] == filters.NONE_TOKEN:
                assert v["label"].startswith("(")
                assert filters.NONE_TOKEN not in v["label"]
    # the sentinel still round-trips through the URL, which is how it is applied
    assert filters.NONE_TOKEN in flt.query_dict()["term"]


def test_channel_comparison_is_volume_gated(ft_db):
    program = programs.get("ft")
    res = _db_matrix(ft_db, "ft")
    comp = metrics.channel_comparison(program, res, min_n=25, stage="admitted")
    assert all(r["n"] >= 25 for r in comp["rows"])
    assert all(r["channel"] != taxonomy.NO_UTM for r in comp["rows"])
    # sorted best-converting first
    rates = [r["rate"] for r in comp["rows"]]
    assert rates == sorted(rates, reverse=True)
    assert abs(comp["baseline_rate"] - 350 / 8436) < 1e-12


def test_slice_selection_is_rank_ordered_and_capped():
    """The rule app.js mirrors. If this changes, the JS twin must change with it:
    which eight series survive, and therefore their colours, is decided here."""
    names = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j"]
    # Click order must not matter — canonical rank does.
    kept, dropped = metrics.slice_selection(names, ["j", "b", "a"])
    assert kept == ["a", "b", "j"]
    assert dropped == []
    kept, dropped = metrics.slice_selection(names, names)
    assert kept == names[:8]
    assert dropped == ["i", "j"]
    assert metrics.slice_selection(names, []) == ([], [])
    # Unknown names are ignored rather than drawn as blanks.
    assert metrics.slice_selection(names, ["zz", "c"]) == (["c"], [])


def test_uncapped_timeline_ships_every_entity_with_a_rank(ft_db):
    """The client payload. Every entity present, each with a stable rank, and
    the first `limit` by rank identical to what the capped call draws."""
    conn, prog = ft_db, programs.PROGRAMS["ft"]
    flt = filters.Filters(prog)
    full = metrics.tag_timeline(conn, prog, flt.where, flt.params,
                                picked=None, cap=False, bucket="month")
    picked = _tl_parents(conn, prog)
    capped = metrics.tag_timeline(conn, prog, flt.where, flt.params,
                                  picked=picked, bucket="month")

    entities = []
    for sv in full["series"]:
        if sv["group"] not in entities:
            entities.append(sv["group"])
    assert len(entities) > 8, "uncapped payload should exceed the palette"
    ranks = {sv["group"]: sv["rank"] for sv in full["series"]}
    assert sorted(ranks.values())[:len(entities)] == list(range(len(entities)))

    by_rank = sorted(entities, key=lambda n: ranks[n])
    kept, _ = metrics.slice_selection(by_rank, picked)
    drawn = {sv["group"] for sv in capped["series"]}
    assert set(kept) == drawn, "client slice must keep exactly what Python draws"

    # And the same hues. Colour is the position within the kept set in taxonomy
    # order — the second half of the rule app.js mirrors. If this drifts, a
    # channel changes colour the moment you tick a box, which is the one thing
    # the palette contract forbids.
    torder = {sv["group"]: sv["torder"] for sv in full["series"]}
    client_colour = {n: i for i, n in
                     enumerate(sorted(kept, key=lambda n: torder[n]))}
    server_colour = {sv["group"]: sv["colour_index"] for sv in capped["series"]}
    assert client_colour == server_colour
    # Same buckets on both sides, or the client would redraw against a
    # different x-axis than the one already on screen.
    assert full["labels"] == capped["labels"]


def test_channel_charts_follow_the_selected_stage(ft_db):
    """Both channel charts answer their question about whichever stage is asked
    for, and neither one changes the population it asks it of."""
    conn, prog = ft_db, programs.PROGRAMS["ft"]
    flt = filters.Filters(prog)
    apps, flags = metrics.load_population(conn, prog, flt.where, flt.params)
    pings = metrics.load_pings(conn, [a["id"] for a in apps])
    m = metrics.build_matrix(prog, apps, flags, pings, "any")

    # Default is still the last stage — the selector must not move the baseline
    # view anyone already has bookmarked.
    assert (metrics.channel_comparison(prog, m)["final_label"]
            == prog.stage_labels[prog.channel_stage])

    for stage in prog.stage_keys[1:]:
        cmp_ = metrics.channel_comparison(prog, m, stage=stage)
        mk = metrics.channel_makeup(prog, m, stage=stage)
        assert cmp_["final_label"] == prog.stage_labels[stage]
        assert mk["final_label"] == prog.stage_labels[stage]
        # Presence denominator is the stage total, not the population.
        assert mk["final_total"] == m["totals"][stage]
        # Quality is reached/touched, and the dashed baseline is the same ratio
        # over everyone — both must move to the SAME stage, or the line would
        # answer a different question than the bars.
        assert cmp_["baseline_rate"] == pytest.approx(
            m["totals"][stage] / m["totals"]["started"], rel=1e-9)
        for r in cmp_["rows"]:
            assert r["rate"] == pytest.approx(r["final_n"] / r["n"], rel=1e-9)
        for r in mk["rows"]:
            assert r["share"] == pytest.approx(
                r["final_n"] / m["totals"][stage], rel=1e-9)

    # Rates rise as the target stage gets easier to reach.
    early = metrics.channel_comparison(prog, m, stage="submitted")
    late = metrics.channel_comparison(prog, m, stage=prog.channel_stage)
    assert early["baseline_rate"] > late["baseline_rate"]


def test_stage_selector_does_not_collide_with_the_reached_filter():
    """`cvs`/`mvs` pick what each channel chart measures; `stage` filters who is
    in scope. They were briefly the same param, which quietly narrowed the
    population to people who had already reached the stage — so every channel
    converted at 100%."""
    from conftest import APP_ROOT
    src = open(os.path.join(APP_ROOT, "app", "main.py"), encoding="utf-8").read()
    assert '_stage("cvs"), _stage("mvs")' in src
    assert 'q.get("stage")' not in src
    # And the filter layer still owns `stage`.
    assert "stage" in filters.DIMENSION_PARAMS["stage"]


def test_each_channel_card_keeps_its_own_stage(ft_db):
    """The two cards are independent: setting one must not move the other, and
    neither may touch anything else on the page."""
    conn, prog = ft_db, programs.PROGRAMS["ft"]
    flt = filters.Filters(prog)
    apps, flags = metrics.load_population(conn, prog, flt.where, flt.params)
    pings = metrics.load_pings(conn, [a["id"] for a in apps])
    m = metrics.build_matrix(prog, apps, flags, pings, "any")

    last = prog.channel_stage
    early = prog.stage_keys[1]
    assert early != last

    # Quality moved, presence left alone.
    assert (metrics.channel_comparison(prog, m, stage=early)["final_label"]
            == prog.stage_labels[early])
    assert metrics.channel_makeup(prog, m)["final_total"] == m["totals"][last]
    # ...and the reverse.
    assert metrics.channel_makeup(prog, m, stage=early)["final_total"] == m["totals"][early]
    assert (metrics.channel_comparison(prog, m)["final_label"]
            == prog.stage_labels[last])

    # Neither one is allowed to reach the stage-presence chart or the tables,
    # which read every stage from the same matrix regardless.
    pen = metrics.penetration_series(prog, m)
    assert pen["stages"] == [prog.stage_labels[k] for k in prog.stage_keys]
    assert pen["totals"] == [m["totals"][k] for k in prog.stage_keys]


def test_stage_default_is_passed_to_the_client_not_inferred():
    """The client drops the URL param when you pick the default stage. It must
    be TOLD which stage that is: Jinja's |tojson sorts dict keys, so the
    payload's last key is alphabetical ("submitted"), not the funnel's last
    stage — inferring it inverted the rule, and every URL came out wrong."""
    from conftest import APP_ROOT
    tpl = open(os.path.join(APP_ROOT, "app", "templates", "overview.html"),
               encoding="utf-8").read()
    assert "program.channel_stage|tojson" in tpl
    js = open(os.path.join(APP_ROOT, "app", "static", "app.js"),
              encoding="utf-8").read()
    assert "key === defaultStage" in js
    assert "keys[keys.length - 1]" not in js


def test_enrolled_stage_nests_inside_admitted(ft_db):
    """Enrolled is derived here, not by the vendored engine, so it gets its own
    guard. Every enrolled person must also be admitted — otherwise the funnel
    would widen at the bottom and every penetration figure below Admitted would
    be measured against a denominator that doesn't contain its own numerator."""
    prog = programs.get("ft")
    apps, flags = metrics.load_population(ft_db, prog, "program = ?", ["ft"])
    overall = metrics.overall_funnel(prog, flags)

    assert prog.stage_keys[-1] == "enrolled"
    assert prog.stage_labels["enrolled"] == "Enrolled"
    assert overall["counts"]["enrolled"] == 16
    assert overall["counts"]["enrolled"] <= overall["counts"]["admitted"]

    both = sum(1 for f in flags if f["enrolled"] and f["admitted"])
    assert both == overall["counts"]["enrolled"], \
        "an enrolled applicant who is not admitted breaks funnel monotonicity"

    # Summer has no such stage and must not have grown one.
    assert "enrolled" not in programs.get("summer").stage_keys


def test_enrolled_reads_a_real_column_when_one_arrives():
    """Today Enrolled comes off Most Recent Decision. The next export is
    expected to carry a dedicated column; when it does, it must win — that is
    the whole reason the stage is being added before the data changes."""
    prog = programs.get("ft")
    row = [""] * 20
    row[9] = "Enrolled"
    assert prog.stages(row, None)["enrolled"] is True
    row[9] = "Accept"
    assert prog.stages(row, None)["enrolled"] is False

    # With a dedicated column, the decision field is ignored entirely.
    row = [""] * 21
    row[9] = "Accept"          # decision says otherwise...
    row[20] = "Y"              # ...but the column is authoritative
    assert prog.stages(row, 20)["enrolled"] is True
    row[20] = "N"
    assert prog.stages(row, 20)["enrolled"] is False
    row[20] = "2026-09-01"     # a date column counts as set
    assert prog.stages(row, 20)["enrolled"] is True

    # Header detection finds the column without matching "Most Recent Decision".
    assert programs.enrolled_index(["Global ID", "Most Recent Decision"]) is None
    assert programs.enrolled_index(["Admitted", "Enrolled"]) == 1
    assert programs.enrolled_index(["Admitted", "Enrolled Date"]) == 1
    assert programs.enrolled_index(["Admitted", "Matriculated"]) == 1


def test_existing_database_gains_the_new_stage_column(tmp_path):
    """An existing funnel.db predates st_enrolled. CREATE TABLE IF NOT EXISTS
    would leave it alone and every insert would then fail on an unknown
    column, so connect() migrates."""
    path = str(tmp_path / "old.db")
    old = sqlite3.connect(path)
    old.executescript(db.SCHEMA.replace(
        "    st_enrolled     INTEGER NOT NULL DEFAULT 0,\n", ""))
    old.commit()
    cols = {r[1] for r in old.execute("PRAGMA table_info(applicants)")}
    assert "st_enrolled" not in cols
    old.close()

    conn = db.connect(path)
    cols = {r["name"] for r in conn.execute("PRAGMA table_info(applicants)")}
    assert "st_enrolled" in cols
    # And it is usable: a full ingest against the migrated file.
    ingest.ingest(conn, FT_FILE, "ft", "ft.xlsx", "")
    prog = programs.get("ft")
    apps, flags = metrics.load_population(conn, prog, "program = ?", ["ft"])
    assert metrics.overall_funnel(prog, flags)["counts"]["enrolled"] == 16
    conn.close()


# media spend
# --------------------------------------------------------------------------
GOOGLE_SPEND_FILE = os.path.join(
    os.path.dirname(FT_FILE), "Google Media Spend FY2526 monthly.csv")
META_SPEND_FILE = os.path.join(
    os.path.dirname(FT_FILE), "Meta Media Spend FY2526 monthly.xlsx")
GOOGLE_NO_MONTH_FILE = os.path.join(
    os.path.dirname(FT_FILE), "Google Media Spend FY2526 no-month.csv")

GOOGLE_TOTAL = 140222.15
META_TOTAL = 148135.28


def _spend_db():
    conn = db.connect(":memory:")
    ingest.ingest(conn, FT_FILE, "ft", "ft.xlsx", "")
    for path, plat in ((GOOGLE_SPEND_FILE, spend.GOOGLE),
                       (META_SPEND_FILE, spend.META)):
        rows, meta = spend.parse(path, plat, os.path.basename(path))
        spend.store(conn, "ft", rows, meta, "")
    return conn


def test_spend_parsers_total_exactly():
    """The whole feature rests on these two numbers. A dropped preamble line, a
    mis-parsed thousands separator or a skipped month all show up right here."""
    rows, meta = spend.parse_google(GOOGLE_SPEND_FILE, "google.csv")
    assert meta["total_cost"] == pytest.approx(GOOGLE_TOTAL, abs=0.01)
    assert sum(r["cost"] for r in rows) == pytest.approx(GOOGLE_TOTAL, abs=0.01)
    assert len(meta["months"]) == 11
    assert meta["period_start"] == "2025-09-01" and meta["period_end"] == "2026-07-31"
    assert not meta["warnings"]

    rows, meta = spend.parse_meta(META_SPEND_FILE, "meta.xlsx")
    assert meta["total_cost"] == pytest.approx(META_TOTAL, abs=0.01)
    assert sum(r["cost"] for r in rows) == pytest.approx(META_TOTAL, abs=0.01)
    assert len(meta["months"]) == 9
    assert not meta["warnings"]


def test_spend_resolves_onto_the_real_taxonomy():
    """Spend joins to the funnel BY NAME. A sub-source that taxonomy.py does not
    know would orphan every dollar in it silently, so the parser refuses."""
    known = {ch: set(subs) for ch, subs in taxonomy.TAXONOMY}
    for path, plat in ((GOOGLE_SPEND_FILE, spend.GOOGLE),
                       (META_SPEND_FILE, spend.META)):
        rows, _ = spend.parse(path, plat, "x")
        for r in rows:
            assert r["sub_source"] in known[r["channel"]]
    google, _ = spend.parse_google(GOOGLE_SPEND_FILE, "g")
    subs = {r["sub_source"] for r in google}
    assert subs == {"PMax", "Search"}
    meta, _ = spend.parse_meta(META_SPEND_FILE, "m")
    assert {"Instagram", "Facebook"} <= {r["sub_source"] for r in meta}


def test_reupload_replaces_rather_than_accumulates():
    """Spend is a restatement of a period, not an event log — platforms revise
    figures after the fact. Ingesting the same file twice must not double it."""
    conn = db.connect(":memory:")
    rows, meta = spend.parse_google(GOOGLE_SPEND_FILE, "g.csv")
    spend.store(conn, "ft", rows, meta, "")
    first = conn.execute("SELECT SUM(cost) FROM spend WHERE program='ft'").fetchone()[0]
    res = spend.store(conn, "ft", rows, meta, "")
    second = conn.execute("SELECT SUM(cost) FROM spend WHERE program='ft'").fetchone()[0]
    assert first == pytest.approx(GOOGLE_TOTAL, abs=0.01)
    assert second == pytest.approx(first, abs=0.01)
    assert res["replaced"] > 0, "the second upload should report what it displaced"
    conn.close()


def test_month_less_export_still_imports():
    """An export pulled without a month segment must still load — into a single
    bucket — and say so, rather than being rejected."""
    rows, meta = spend.parse_google(GOOGLE_NO_MONTH_FILE, "nomonth.csv")
    assert meta["months"] == []
    assert all(r["month"] == "" for r in rows)
    assert any("month" in w.lower() for w in meta["warnings"])
    conn = db.connect(":memory:")
    spend.store(conn, "ft", rows, meta, "")
    assert conn.execute("SELECT COUNT(*) FROM spend").fetchone()[0] == len(rows)
    conn.close()


def test_cost_columns_that_sum_do_and_the_one_that_cannot_is_flagged():
    """The trap this view exists to avoid. First-touch denominators partition
    people and therefore sum; any-touch ones overlap and must not. A blended
    cost-per-admit built on any-touch counts read 44% too cheap."""
    conn = _spend_db()
    prog = programs.get("ft")
    flt = filters.Filters(prog)
    apps, flags = metrics.load_population(conn, prog, flt.where, flt.params)
    pings = metrics.load_pings(conn, [a["id"] for a in apps])
    anym = metrics.build_matrix(prog, apps, flags, pings, "any")
    ftm = metrics.build_matrix(prog, apps, flags, pings, "first")
    cost = metrics.cost_by_channel(conn, prog, anym, ftm, "admitted")

    parents = [r for r in cost["rows"]]
    assert cost["total_cost"] == pytest.approx(GOOGLE_TOTAL + META_TOTAL, abs=0.02)
    assert sum(r["cost"] for r in parents) == pytest.approx(cost["total_cost"], abs=0.02)

    # first touch: the parts equal the whole
    assert sum(r["first_n"] for r in parents) == cost["total_first"]
    assert sum(r["stage_n"] for r in parents) == cost["stage_total"]
    assert cost["blended_per_start"] == pytest.approx(
        cost["total_cost"] / cost["total_first"])

    # any touch: strictly larger, because people touch more than one channel
    assert sum(r["n"] for r in parents) > cost["total_first"]
    assert sum(r["any_stage_n"] for r in parents) > cost["stage_total"]

    # sub-sources roll up to their parent's spend
    for r in parents:
        if r["subs"]:
            assert sum(s["cost"] for s in r["subs"]) == pytest.approx(r["cost"], abs=0.02)
    conn.close()


def test_spend_with_no_funnel_rows_still_appears():
    """A channel burning money with nothing to show is the most useful thing
    this table can say. Dropping the row would hide it."""
    conn = _spend_db()
    prog = programs.get("ft")
    flt = filters.Filters(prog)
    apps, flags = metrics.load_population(conn, prog, flt.where, flt.params)
    pings = metrics.load_pings(conn, [a["id"] for a in apps])
    cost = metrics.cost_by_channel(
        conn, prog,
        metrics.build_matrix(prog, apps, flags, pings, "any"),
        metrics.build_matrix(prog, apps, flags, pings, "first"), "admitted")
    meta_row = next(r for r in cost["rows"] if r["name"] == spend.META_CHANNEL)
    other = next(s for s in meta_row["subs"] if s["name"] == "Other Meta")
    assert other["cost"] > 0 and other["first_n"] == 0
    assert other["cost_per_start"] is None, "no starts must read as a dash, not $0"
    conn.close()


def test_meta_ad_id_coverage_proves_the_files_describe_the_same_traffic():
    """Meta tags links utm_content={{ad.id}}, so spend can be checked ad by ad
    against the ping log. Nothing in the rollup depends on it; it is the signal
    that says the spend file and the funnel have not drifted apart."""
    conn = _spend_db()
    cov = spend.meta_ad_coverage(conn, "ft")
    assert cov["spend_pct"] > 0.95
    assert cov["pings_pct"] > 0.85
    assert cov["matched_ads"] > 50
    assert spend.meta_ad_coverage(conn, "summer") is None
    conn.close()


def test_date_filter_narrows_spend_by_whole_months():
    """Spend is monthly. A range inside one month counts that month WHOLE —
    splitting it would be inventing daily figures — and the caller is told."""
    conn = _spend_db()
    prog = programs.get("ft")
    months, all_months, narrowed = metrics.spend_months(conn, prog, filters.Filters(prog))
    assert months == all_months and not narrowed

    tight = filters.Filters(prog, date_field="started_date",
                            date_from="2026-01-10", date_to="2026-01-20")
    months, all_months, narrowed = metrics.spend_months(conn, prog, tight)
    assert months == ["2026-01"] and narrowed
    conn.close()


def test_date_range_defaults_to_the_current_fiscal_year(ft_db):
    """Every figure here is read for one intake year; the unfiltered view mixes
    several. The default is applied when no date params are present, and must be
    escapable — otherwise it is a hidden filter rather than a starting point."""
    prog = programs.get("ft")

    class Q(dict):
        def getlist(self, k):
            v = self.get(k)
            return [v] if isinstance(v, str) else (v or [])

    fy = filters.default_fiscal_year(ft_db, prog)
    assert fy == 2025
    assert filters.fiscal_range(2025) == ("2025-09-01", "2026-08-31")

    d = filters.from_query(prog, Q(), default_fy=fy)
    assert (d.date_field, d.date_from, d.date_to) == \
        ("started_date", "2025-09-01", "2026-08-31")

    # ...and every escape hatch works.
    allt = filters.from_query(prog, Q(dates="all"), default_fy=fy)
    assert allt.date_field is None and not allt.active

    prev = filters.from_query(prog, Q(dates="2024"), default_fy=fy)
    assert (prev.date_from, prev.date_to) == ("2024-09-01", "2025-08-31")

    custom = filters.from_query(
        prog, Q(date_field="started_date", date_from="2026-01-01",
                date_to="2026-03-31"), default_fy=fy)
    assert (custom.date_from, custom.date_to) == ("2026-01-01", "2026-03-31")

    # The default narrows the population but must not empty it.
    n_all = len(metrics.load_population(ft_db, prog, allt.where, allt.params)[0])
    n_def = len(metrics.load_population(ft_db, prog, d.where, d.params)[0])
    assert n_all == 8436
    assert 0 < n_def < n_all


def test_spend_trend_carries_sub_sources_with_stable_colours(ft_db):
    """The trend picker offers channels AND sub-sources, and a channel keeps its
    hue whichever tab it is drawn on — colour follows the entity, never rank."""
    conn = ft_db
    for path, plat in ((GOOGLE_SPEND_FILE, spend.GOOGLE),
                       (META_SPEND_FILE, spend.META)):
        rows, meta = spend.parse(path, plat, os.path.basename(path))
        spend.store(conn, "ft", rows, meta, "")
    prog = programs.get("ft")

    tree = metrics.spend_tree(conn, prog)
    names = metrics.tree_names(tree)
    assert "Google (Paid)" in names
    assert "Meta Paid Social (IG/FB) " + chr(0x203A) + " Instagram" in names

    full = metrics.spend_trend(conn, prog, cap=False)
    by = {s["name"]: s for s in full["series"]}
    # A parent is the sum of its own sub-sources here. That is true for money
    # and false for any-touch people, which is why the two use different code.
    for parent in ("Google (Paid)", "Meta Paid Social (IG/FB)"):
        kids = [s for n, s in by.items()
                if n.startswith(parent + taxonomy.SUB_SEP)]
        for i in range(len(full["labels"])):
            assert by[parent]["data"][i] == pytest.approx(
                sum(k["data"][i] for k in kids), abs=0.01)

    # Default selection is parents only — switching every sub-source on at once
    # would blow past the eight-colour palette.
    default = metrics.spend_trend(conn, prog)
    assert all(not s["is_sub"] for s in default["series"])


def test_first_and_any_touch_cost_answer_different_questions(ft_db):
    """Both lenses, and the rule that separates them: first touch partitions
    people so its rows sum; any touch overlaps, so its blended figure must come
    from a SET UNION. Summing the any-touch column charges anyone who touched
    both Google and Meta twice."""
    conn = ft_db
    for path, plat in ((GOOGLE_SPEND_FILE, spend.GOOGLE),
                       (META_SPEND_FILE, spend.META)):
        rows, meta = spend.parse(path, plat, os.path.basename(path))
        spend.store(conn, "ft", rows, meta, "")

    prog = programs.get("ft")
    flt = filters.Filters(prog)
    apps, flags = metrics.load_population(conn, prog, flt.where, flt.params)
    pings = metrics.load_pings(conn, [a["id"] for a in apps])
    anym = metrics.build_matrix(prog, apps, flags, pings, "any")
    ftm = metrics.build_matrix(prog, apps, flags, pings, "first")
    paid = [spend.GOOGLE_CHANNEL, spend.META_CHANNEL]
    reach = metrics.paid_reach(prog, apps, flags, pings, paid)

    first = metrics.cost_by_channel(conn, prog, anym, ftm, "admitted",
                                    attribution="first", reach=reach)
    any_ = metrics.cost_by_channel(conn, prog, anym, ftm, "admitted",
                                   attribution="any", reach=reach)

    assert first["rows_sum"] is True and any_["rows_sum"] is False
    assert first["total_cost"] == pytest.approx(any_["total_cost"], abs=0.02)

    # first touch: the parts equal the whole
    assert sum(r["stage_n"] for r in first["rows"]) == first["stage_total"]
    # any touch: the parts EXCEED the whole, and the whole is the union
    assert sum(r["stage_n"] for r in any_["rows"]) > any_["stage_total"]
    assert any_["stage_total"] == reach["admitted"]
    assert any_["total_first"] == reach["_touched"]

    # A channel credited with more people costs less per person.
    for name in (spend.GOOGLE_CHANNEL, spend.META_CHANNEL):
        f = next(r for r in first["rows"] if r["name"] == name)
        a = next(r for r in any_["rows"] if r["name"] == name)
        assert a["stage_n"] >= f["stage_n"]
        assert a["cost_per_stage"] <= f["cost_per_stage"]

    # The union is genuinely smaller than the sum here — people really do touch
    # both. If this ever stops being true the union has quietly become a no-op.
    assert reach["admitted"] < sum(r["stage_n"] for r in any_["rows"])


def test_cost_payload_ships_every_lens_for_every_stage(ft_db):
    """The toggle and the stage chips both redraw client-side, so the payload
    has to carry stages x attributions with no extra queries."""
    conn = ft_db
    rows, meta = spend.parse(GOOGLE_SPEND_FILE, spend.GOOGLE, "g.csv")
    spend.store(conn, "ft", rows, meta, "")
    prog = programs.get("ft")
    flt = filters.Filters(prog)
    apps, flags = metrics.load_population(conn, prog, flt.where, flt.params)
    pings = metrics.load_pings(conn, [a["id"] for a in apps])
    payload = metrics.cost_stage_payload(
        prog and conn, prog,
        metrics.build_matrix(prog, apps, flags, pings, "any"),
        metrics.build_matrix(prog, apps, flags, pings, "first"),
        prog.stage_keys,
        reach=metrics.paid_reach(prog, apps, flags, pings,
                                 [spend.GOOGLE_CHANNEL]))
    assert set(payload) == {"first", "last", "any"}
    for attr in payload:
        assert set(payload[attr]) == set(prog.stage_keys)
        for k, v in payload[attr].items():
            # Only any-touch overlaps; first and last each put a person in
            # exactly one channel, so both sum.
            assert v["rows_sum"] == (attr != "any")
            assert v["stage_label"] == prog.stage_labels[k]


def test_new_slate_layout_with_referral_info_column(tmp_path):
    """2026-08-03: Slate inserted "Referral Info" at index 19, pushing Program
    Emphasis to 20. Both arrangements must ingest, and emphasis must be read
    from the right column in each -- reading index 19 on the new file would
    silently fill emphasis with referrer URLs."""
    import io
    import openpyxl as _ox
    prog = programs.get("ft")

    src = _ox.load_workbook(FT_FILE, read_only=True, data_only=True)
    rows = [list(r) for i, r in
            enumerate(src["Export"].iter_rows(values_only=True)) if i <= 60]
    src.close()

    out = _ox.Workbook()
    ws = out.active
    ws.title = "Export"
    for i, row in enumerate(rows):
        row = list(row)
        row.insert(19, "Referral Info" if i == 0
                   else "2025-09-01 00:00:00 | https://www.aada.edu/")
        ws.append(row)
    buf = io.BytesIO()
    out.save(buf)
    buf.seek(0)

    conn = db.connect(":memory:")
    res = ingest.ingest(conn, buf, "ft", "new-layout.xlsx", "")
    assert res["applicants_new"] > 0

    # Emphasis came from column 20, not the referrer text now sitting at 19.
    vals = {r["emphasis"] for r in
            conn.execute("SELECT DISTINCT emphasis FROM applicants")}
    assert not any("http" in (v or "") for v in vals), \
        "emphasis was read from the Referral Info column"

    # Both layouts resolve, and to different maps.
    old_headers = [c for c in rows[0]]
    new_headers = old_headers[:19] + ["Referral Info"] + old_headers[19:]
    old_layout, _ = prog.layout_for(old_headers)
    new_layout, _ = prog.layout_for(new_headers)
    assert old_layout is not None and new_layout is not None
    assert old_layout.cols["emphasis"] == 19
    assert new_layout.cols["emphasis"] == 20
    conn.close()


def test_upload_kind_is_detected_from_headers_not_filenames():
    """Auto-detect reads column headers, so a badly-named file still routes
    correctly. A filename-only guess is the fallback, never the first answer."""
    import io
    import openpyxl as _ox

    def book(headers):
        out = _ox.Workbook()
        ws = out.active
        ws.title = "Export"
        ws.append(headers)
        ws.append(["x"] * len(headers))
        buf = io.BytesIO()
        out.save(buf)
        return buf.getvalue()

    ft = book(["Global ID", "Full-Time App Term", "Started FT App Date"])
    kind, conf, _why = ingest.sniff_kind("totally-misleading-name.xlsx", ft)
    assert (kind, conf) == (ingest.KIND_FT, "sure")

    summer = book(["Global ID", "Term", "App Date", "Application Status"])
    kind, conf, _why = ingest.sniff_kind("also-wrong.xlsx", summer)
    assert (kind, conf) == (ingest.KIND_SUMMER, "sure")

    meta = book(["Campaign name", "Ad set name", "Amount spent (USD)"])
    kind, conf, _why = ingest.sniff_kind("q3.xlsx", meta)
    assert (kind, conf) == (ingest.KIND_META, "sure")

    google = b"Media Spend\n\"Sept 1 - Jul 31\"\nCampaign type,Campaign,Cost\n"
    kind, conf, _why = ingest.sniff_kind("export.csv", google)
    assert (kind, conf) == (ingest.KIND_GOOGLE, "sure")

    # Unreadable file -> falls back to the name rather than raising.
    kind, conf, _why = ingest.sniff_kind("Ping Data - Summer.xlsx", b"not a workbook")
    assert (kind, conf) == (ingest.KIND_SUMMER, "guess")


def test_last_touch_partitions_people_like_first_touch(ft_db):
    """Last touch answers "what closed them" where first touch answers "what
    found them". Both put a person in exactly ONE channel, so both sum -- only
    any-touch overlaps. Getting this wrong would let the last-touch column be
    silently treated as un-summable, or worse, let any-touch be summed."""
    conn = ft_db
    for path, plat in ((GOOGLE_SPEND_FILE, spend.GOOGLE),
                       (META_SPEND_FILE, spend.META)):
        rows, meta = spend.parse(path, plat, os.path.basename(path))
        spend.store(conn, "ft", rows, meta, "")

    prog = programs.get("ft")
    flt = filters.Filters(prog)
    apps, flags = metrics.load_population(conn, prog, flt.where, flt.params)
    pings = metrics.load_pings(conn, [a["id"] for a in apps])
    M = {t: metrics.build_matrix(prog, apps, flags, pings, t)
         for t in ("any", "first", "last")}
    reach = metrics.paid_reach(prog, apps, flags, pings,
                               [spend.GOOGLE_CHANNEL, spend.META_CHANNEL])

    def cost(attr):
        return metrics.cost_by_channel(
            conn, prog, M["any"], M["first"], "admitted", attribution=attr,
            reach=reach, last_matrix=M["last"])

    first, last, any_ = cost("first"), cost("last"), cost("any")

    assert last["rows_sum"] is True and first["rows_sum"] is True
    assert any_["rows_sum"] is False

    # Last touch sums, exactly like first touch.
    assert sum(r["stage_n"] for r in last["rows"]) == last["stage_total"]
    # ...and is a genuinely different answer, not a copy of first touch.
    assert last["stage_total"] != first["stage_total"]

    # Every lens divides the SAME spend -- only the denominator moves.
    assert (last["total_cost"] == pytest.approx(first["total_cost"], abs=0.02)
            == pytest.approx(any_["total_cost"], abs=0.02))

    # Any touch credits the most people, so it is always the cheapest lens.
    assert any_["stage_total"] > first["stage_total"]
    assert any_["stage_total"] > last["stage_total"]
    assert any_["blended_per_stage"] < first["blended_per_stage"]
    assert any_["blended_per_stage"] < last["blended_per_stage"]


def test_shrinking_term_warns_but_never_deletes(ft_db):
    """Every export is expected to be a superset of the last one. A term coming
    back with fewer rows than the database already holds must be surfaced loud
    — this is exactly what happened when Slate's Aug 2026 pull dropped 1,326 of
    1,334 Winter 2026 rows — but must never delete anything: uploads only add
    or update."""
    import io
    import openpyxl as _ox

    prog = programs.get("ft")
    before = conn_count = ft_db.execute(
        "SELECT COUNT(*) FROM applicants WHERE program='ft'").fetchone()[0]
    term_before = dict(ft_db.execute(
        "SELECT term, COUNT(*) FROM applicants WHERE program='ft' GROUP BY term"
    ).fetchall())
    biggest_term = max(term_before, key=term_before.get)
    assert term_before[biggest_term] > 10

    # Rebuild a small file containing only rows NOT in that term -- i.e. this
    # term drops to zero, everything else is absent too (a much narrower pull).
    src = _ox.load_workbook(FT_FILE, read_only=True, data_only=True)
    all_rows = list(src["Export"].iter_rows(values_only=True))
    src.close()
    header, data_rows = all_rows[0], all_rows[1:]
    narrow = [r for r in data_rows if r[1] != biggest_term][:50]

    out = _ox.Workbook()
    ws = out.active
    ws.title = "Export"
    ws.append(list(header))
    for r in narrow:
        ws.append(list(r))
    buf = io.BytesIO()
    out.save(buf)
    buf.seek(0)

    res = ingest.ingest(ft_db, buf, "ft", "narrower.xlsx", "")
    warn = next((w for w in res["warnings"] if "FEWER rows" in w), None)
    assert warn is not None
    assert repr(biggest_term) in warn or (biggest_term or "(blank)") in warn

    after = ft_db.execute(
        "SELECT COUNT(*) FROM applicants WHERE program='ft'").fetchone()[0]
    assert after >= before, "a narrower file must never reduce what is stored"
    after_term = ft_db.execute(
        "SELECT COUNT(*) FROM applicants WHERE program='ft' AND term=?",
        (biggest_term,)).fetchone()[0]
    assert after_term == term_before[biggest_term], \
        "the shrunk term's existing rows must survive untouched"


def test_winter_and_spring_are_one_intake():
    """AADA is renaming the January intake from "Winter" to "Spring", and
    Slate's Aug 2026 export carries both spellings at once. Term is part of the
    applicant dedup key, so leaving them distinct would split one intake into
    two filter options AND import the same person twice."""
    assert programs.canonical_term("Winter 2027 (January 2027)") == \
        programs.canonical_term("January 2027 (Spring)") == "January 2027 (Spring)"
    assert programs.canonical_term("Winter 2026 (January 2026)") == "January 2026 (Spring)"

    # The January year wins, not the label year -- they can disagree.
    assert programs.canonical_term("Winter 2025 (January 2025)") == "January 2025 (Spring)"

    # Everything else is left completely alone.
    for untouched in ("Fall 2026", "Summer 2026", "Fall 2025 (August 2025)", ""):
        assert programs.canonical_term(untouched) == untouched
    assert programs.canonical_term(None) == ""


def test_stored_winter_terms_are_migrated_on_connect(tmp_path):
    """A database written before the rename still holds "Winter ..." rows.
    connect() rewrites them, so a later upload of the same people under the new
    label updates those rows instead of inserting duplicates."""
    path = str(tmp_path / "old-terms.db")
    conn = db.connect(path)
    ingest.ingest(conn, FT_FILE, "ft", "seed.xlsx", "")
    # Force a row back to the pre-rename spelling, as an older DB would have.
    conn.execute("UPDATE applicants SET term='Winter 2026 (January 2026)' "
                 "WHERE program='ft' AND term='January 2026 (Spring)'")
    conn.commit()
    stale = conn.execute(
        "SELECT COUNT(*) FROM applicants WHERE term LIKE 'Winter%'").fetchone()[0]
    assert stale > 0
    conn.close()

    conn = db.connect(path)          # migration runs here
    assert conn.execute(
        "SELECT COUNT(*) FROM applicants WHERE term LIKE 'Winter%'").fetchone()[0] == 0
    assert conn.execute(
        "SELECT COUNT(*) FROM applicants WHERE term='January 2026 (Spring)'"
    ).fetchone()[0] == stale
    conn.close()


def test_one_channel_wears_one_colour_everywhere(ft_db):
    """The whole point of the canonical slot map: Meta is the same green on the
    timeline, the stage-presence chart, the first-touch donut and the spend
    trend. Colour identifies the CHANNEL; it must never depend on which series
    happen to be on screen, or on rank within the current view."""
    conn = ft_db
    for path, plat in ((GOOGLE_SPEND_FILE, spend.GOOGLE),
                       (META_SPEND_FILE, spend.META)):
        rows, meta = spend.parse(path, plat, os.path.basename(path))
        spend.store(conn, "ft", rows, meta, "")

    prog = programs.get("ft")
    flt = filters.Filters(prog)
    apps, flags = metrics.load_population(conn, prog, flt.where, flt.params)
    pings = metrics.load_pings(conn, [a["id"] for a in apps])
    anym = metrics.build_matrix(prog, apps, flags, pings, "any")

    META = spend.META_CHANNEL
    want = taxonomy.channel_slot(META)
    assert want is not None

    tl = metrics.tag_timeline(conn, prog, flt.where, flt.params, bucket="month")
    tl_slots = {s["colour_index"] for s in tl["series"] if s["group"] == META}
    assert tl_slots == {want}

    pen = metrics.penetration_series(prog, anym)
    pen_slots = {s["colour_index"] for s in pen["series"] if s["name"] == META}
    assert pen_slots == {want}

    trend = metrics.spend_trend(conn, prog, cap=False)
    tr_slots = {s["colour_index"] for s in trend["series"] if s["name"] == META}
    assert tr_slots == {want}

    # ...and it does not move when the selection changes around it.
    narrow = metrics.tag_timeline(conn, prog, flt.where, flt.params,
                                  bucket="month", picked=[META])
    assert {s["colour_index"] for s in narrow["series"]} == {want}

    # A sub-source shares its parent's slot, by design.
    sub = taxonomy.sub_name(META, "Instagram")
    assert taxonomy.channel_slot(sub) == want

    # Channels past the 8 validated hues get no slot rather than reusing one.
    assert taxonomy.channel_slot("Spotify (Paid Audio)") is None
    majors = [c for c, _s in taxonomy.TAXONOMY[:8]]
    slots = [taxonomy.channel_slot(c) for c in majors]
    assert sorted(slots) == list(range(8)), "the 8 majors must not collide"
